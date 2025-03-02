from collections import defaultdict
from datetime import datetime, timedelta
from copy import deepcopy
import calendar
import os
import re
import io
import calendar
import logging
import pandas as pd
import time
from sqlalchemy.exc import OperationalError
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from sqlalchemy import and_, or_
from logging.handlers import RotatingFileHandler
from decimal import Decimal, getcontext, ROUND_HALF_UP
from sqlalchemy import create_engine, Column, String, Float, DateTime, Integer, ForeignKey, func, text
from sqlalchemy.orm import declarative_base, sessionmaker, scoped_session, relationship
from telegram import Update
from decimal import Decimal, ROUND_HALF_UP
from sqlalchemy import Numeric
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes
)
from PIL import Image, ImageDraw, ImageFont

# 修正后的样式定义
HEADER_FILL = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # 灰色标题行
HIGHLIGHT_FILL = PatternFill(start_color='B4C6E7', fill_type='solid')  # 浅蓝色（期初/期末行）
POSITIVE_FONT = Font(color='008000')  # 绿色（正数）
NEGATIVE_FONT = Font(color='FF0000')  # 红色（负数）
CUSTOMER_PAYMENT_FILL = PatternFill(start_color='C6EFCE', fill_type='solid')  # 浅绿（客户支付）
COMPANY_PAYMENT_FILL = PatternFill(start_color='FFC7CE', fill_type='solid')  # 浅红（公司支付）

# Define colors to match Excel
header_color = '#B8CCE4'  # Light blue for header
border_color = '#000000'
alt_row_color = '#FFFFFF'  # White for alternating rows
initial_balance_color = '#B4C6E7'  # Light blue for initial balance
final_balance_color = '#FFFF00'  # Bright yellow for final balance
positive_color = '#008000'  # Green for positive values
negative_color = '#FF0000'  # Red for negative values
customer_payment_color = '#C6EFCE'  # Light green for customer payments
company_payment_color = '#FFC7CE'  # Light pink for company payments

# ================== 初始化配置 ==================
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)
getcontext().prec = 8
Base = declarative_base()

# ================== 数据库模型 ==================
class Customer(Base):
    __tablename__ = 'customers'
    name = Column(String(50), primary_key=True)
    wallet = Column(String(34))
    balances = relationship("Balance", back_populates="customer")

class Balance(Base):
    __tablename__ = 'balances'
    id = Column(Integer, primary_key=True)
    customer_name = Column(String(50), ForeignKey('customers.name'))
    currency = Column(String(4))
    amount = Column(Float)
    customer = relationship("Customer", back_populates="balances")

class Transaction(Base):
    __tablename__ = 'transactions'
    order_id = Column(String(12), primary_key=True)
    customer_name = Column(String(50))
    transaction_type = Column(String(10))  # 'buy', 'sell', 或 'payment'
    sub_type = Column(String(20))  # 例如 '客户支付' 或 '公司支付'
    base_currency = Column(String(4))
    quote_currency = Column(String(4))
    amount = Column(Float)
    rate = Column(Float)
    operator = Column(String(1))
    status = Column(String(20), default='pending')
    payment_in = Column(Float, default=0)
    payment_out = Column(Float, default=0)
    timestamp = Column(DateTime, default=datetime.now)
    settled_in = Column(Float, default=0)  
    settled_out = Column(Float, default=0)

class Adjustment(Base):
    __tablename__ = 'adjustments'
    id = Column(Integer, primary_key=True)
    customer_name = Column(String(50))
    currency = Column(String(4))
    amount = Column(Float)
    note = Column(String(200))
    timestamp = Column(DateTime, default=datetime.now)

class Expense(Base):
    __tablename__ = 'expenses'
    id = Column(Integer, primary_key=True)
    amount = Column(Float)
    currency = Column(String(4))
    purpose = Column(String(200))
    timestamp = Column(DateTime, default=datetime.now)

class USDTAverageCost(Base):
    __tablename__ = 'usdt_average_cost'
    id = Column(Integer, primary_key=True)
    total_usdt = Column(Float, default=0.0)       # 累计获得USDT总量
    total_myr_spent = Column(Float, default=0.0)  # 累计消耗的MYR
    average_cost = Column(Float, default=0.0)     # 平均成本（MYR/USDT）

class MYRAverageCost(Base):
    __tablename__ = 'myr_average_cost'
    id = Column(Integer, primary_key=True)
    total_myr = Column(Float, default=0.0)        # 累计获得MYR总量
    total_usdt_spent = Column(Float, default=0.0) # 累计消耗的USDT
    average_cost = Column(Float, default=0.0)     # 平均成本（USDT/MYR）

# ================== 数据库初始化 ==================
engine = create_engine('sqlite:///fx_bot.db', pool_pre_ping=True, connect_args={'timeout': 30})
Base.metadata.create_all(engine)
session_factory = sessionmaker(bind=engine)
Session = scoped_session(session_factory)


# ================== 数据库迁移脚本 ==================
def run_migrations():
    engine = create_engine('sqlite:///fx_bot.db')
    with engine.connect() as conn:
        try:
            conn.execute(text("ALTER TABLE transactions ADD COLUMN settled_in FLOAT DEFAULT 0"))
            conn.execute(text("ALTER TABLE transactions ADD COLUMN settled_out FLOAT DEFAULT 0"))
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS usdt_average_cost (
                    id INTEGER PRIMARY KEY,
                    total_usdt FLOAT DEFAULT 0.0,
                    total_myr_spent FLOAT DEFAULT 0.0,
                    average_cost FLOAT DEFAULT 0.0
                )
            """))
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS myr_average_cost (
                    id INTEGER PRIMARY KEY,
                    total_myr FLOAT DEFAULT 0.0,
                    total_usdt_spent FLOAT DEFAULT 0.0,
                    average_cost FLOAT DEFAULT 0.0
                )
            """))
            conn.commit()
            logger.info("新增均价表迁移成功")
        except Exception as e:
            logger.warning("数据库迁移可能已经完成: %s", str(e))

def initialize_average_cost(session):
    if not session.query(USDTAverageCost).first():
        session.add(USDTAverageCost())
    if not session.query(MYRAverageCost).first():
        session.add(MYRAverageCost())
    session.commit()

# 在应用启动时调用
initialize_average_cost(Session())

# ================== 核心工具函数 ==================
def setup_logging():
    """配置日志系统"""
    log_dir = "logs"
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, "fx_bot.log")
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            RotatingFileHandler(log_file, maxBytes=5*1024*1024, backupCount=3),
            logging.StreamHandler()
        ]
    )
    logger.info("日志系统初始化完成")

def generate_order_id(session):
    """生成递增订单号，只考虑以 'YS' 开头的订单号"""
    last_order = session.query(Transaction)\
        .filter(Transaction.order_id.like("YS%"))\
        .order_by(Transaction.order_id.desc()).first()
    if last_order:
        last_num = int(last_order.order_id[2:])
        return f"YS{last_num + 1:09d}"
    return "YS000000001"


def generate_payment_id(session, prefix):
    """
    生成一个唯一的支付记录订单号，格式例如 "PAY-R-<timestamp>-<random>"。
    使用 session 来检查是否已存在相同订单号。
    """
    while True:
        new_id = f"{prefix}-{int(time.time())}-{random.randint(1000, 9999)}"
        # 使用 session.no_autoflush 避免自动flush造成问题
        with session.no_autoflush:
            existing = session.query(Transaction).filter(Transaction.order_id == new_id).first()
        if not existing:
            return new_id
        
def update_balance(session, customer: str, currency: str, amount: float):
    """安全的余额更新（支持4位货币代码）"""
    try:
        # 确保客户记录存在
        customer_obj = session.query(Customer).filter_by(name=customer).first()
        if not customer_obj:
            customer_obj = Customer(name=customer)
            session.add(customer_obj)
            session.flush()  # 立即写入数据库但不提交事务

        currency = currency.upper()  # 移除截断，保留完整货币代码
        balance = session.query(Balance).filter_by(
            customer_name=customer,
            currency=currency
        ).with_for_update().first()

        new_amount = round(amount, 2)
        if balance:
            balance.amount = round(balance.amount + new_amount, 2)
        else:
            balance = Balance(
                customer_name=customer,
                currency=currency,
                amount=new_amount
            )
            session.add(balance)
        logger.info(f"余额更新: {customer} {currency} {new_amount:+}")
    except Exception as e:
        logger.error(f"余额更新失败: {str(e)}")
        raise

def get_balance(customer: str, currency: str, session=None) -> float:
    """
    查询指定客户在指定币种下的当前余额。
    如果传入了 session，则使用该 session；否则新建一个 session 查询后关闭。
    """
    # 如果已经传入session则使用它，否则自己创建session
    close_session = False
    if session is None:
        from fx_bot import Session  # 或者你相应的导入路径
        session = Session()
        close_session = True
    try:
        balance_obj = session.query(Balance).filter_by(customer_name=customer, currency=currency.upper()).first()
        return balance_obj.amount if balance_obj else 0.0
    finally:
        if close_session:
            session.close()

def parse_date_range(date_str: str):
    """解析日期范围字符串"""
    try:
        start_str, end_str = date_str.split('-')
        start_date = datetime.strptime(start_str.strip(), '%d/%m/%Y')
        end_date = datetime.strptime(end_str.strip(), '%d/%m/%Y')
        # 将结束日期设置为当天的23:59:59
        end_date = end_date.replace(hour=23, minute=59, second=59)
        return start_date, end_date
    except Exception as e:
        raise ValueError("日期格式错误，请使用 DD/MM/YYYY-DD/MM/YYYY 格式")

STATUS_LIST = ['pending', 'partial', '进行中', '部分结算']

# 辅助函数，判断订单是否完全结清
def is_fully_settled(tx):
    """
    判断订单是否完全结清（即双边都结算完毕）。
    对于买入订单（客户支付报价币，获得基础币）：
      - 客户支付金额（settled_in）应达到：若 operator=='/' 则 amount / rate，否则 amount * rate
      - 公司支付金额（settled_out）应达到订单金额（amount，即基础币数量）
    对于卖出订单（客户支付基础币，获得报价币）：
      - 客户支付金额（settled_in）应达到订单金额（amount，即基础币数量）
      - 公司支付金额（settled_out）应达到：若 operator=='/' 则 amount / rate，否则 amount * rate
    """
    tolerance = Decimal('0.01')
    amount = Decimal(str(tx.amount))
    rate = Decimal(str(tx.rate))
    si = Decimal(str(tx.settled_in or 0))
    so = Decimal(str(tx.settled_out or 0))
    if tx.transaction_type == 'buy':
        if tx.operator == '/':
            expected_in = (amount / rate).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
        else:
            expected_in = (amount * rate).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
        expected_out = amount  # 基础币
        return (si >= expected_in - tolerance) and (so >= expected_out - tolerance)
    elif tx.transaction_type == 'sell':
        if tx.operator == '/':
            expected_out = (amount / rate).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
        else:
            expected_out = (amount * rate).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
        expected_in = amount  # 基础币
        return (si >= expected_in - tolerance) and (so >= expected_out - tolerance)
    else:
        return False
        
def match_buy_order(buy_order, sell_orders, usdt_avg, myr_avg):
    """
    对一个买入订单与多个卖出订单进行匹配，返回：
      - cost_usdt: 累计的USDT成本
      - cost_myr: 累计的MYR成本
      - matched_sell_ids: 匹配到的卖出订单号列表

    参数说明：
      buy_order: 字典，至少包含 'order_id', 'quote_currency', 'amount', 'remaining'
      sell_orders: 卖出订单列表，每个元素为字典，必须包含 'order_id', 'quote_currency',
                   'amount', 'original_cost' 和 'remaining'
      usdt_avg: 当前USDT平均成本（例如：4.42，表示1 USDT = 4.42 MYR）
      myr_avg: 当前MYR平均成本（例如：0.226，表示1 MYR = 0.226 USDT）
    """
    cost_usdt = 0.0
    cost_myr = 0.0
    matched_sell_ids = []
    remaining = buy_order['remaining']

    # 遍历整个卖出订单列表（不使用全局索引）
    for sell in sell_orders:
        if sell['remaining'] <= 1e-6:
            continue  # 已经消耗完的订单跳过

        # 计算本次匹配的金额
        matched = min(remaining, sell['remaining'])
        ratio = matched / sell['amount']  # 计算占用比例
        actual_cost = ratio * sell['original_cost']  # 按比例计算成本

        # 调试日志
        logger.debug(f"匹配买单 {buy_order['order_id']} 与卖单 {sell['order_id']}: "
                     f"matched={matched}, ratio={ratio:.4f}, actual_cost={actual_cost:.4f}")

        # 如果币种一致，直接累加；否则进行跨币种转换
        if buy_order['quote_currency'] == sell['quote_currency']:
            if buy_order['quote_currency'] == 'USDT':
                cost_usdt += actual_cost
            else:
                cost_myr += actual_cost
        else:
            if buy_order['quote_currency'] == 'USDT' and sell['quote_currency'] == 'MYR':
                # 将 MYR 成本转换为 USDT：除以 usdt_avg
                cost_usdt += actual_cost / usdt_avg
            elif buy_order['quote_currency'] == 'MYR' and sell['quote_currency'] == 'USDT':
                # 将 USDT 成本转换为 MYR：除以 myr_avg
                cost_myr += actual_cost / myr_avg

        remaining -= matched
        sell['remaining'] -= matched
        matched_sell_ids.append(sell['order_id'])

        # 调试日志：剩余匹配量
        logger.debug(f"买单 {buy_order['order_id']} 剩余匹配量：{remaining:.4f}")

        if remaining <= 1e-6:
            break

    # 更新买单的 remaining 为匹配后的剩余金额
    buy_order['remaining'] = remaining
    return cost_usdt, cost_myr, matched_sell_ids

def convert_currency(amount, source_currency, target_currency, usdt_avg, myr_avg):
    """
    将金额从source_currency转换为target_currency
    """
    if source_currency == target_currency:
        return amount
    elif source_currency == 'MYR' and target_currency == 'USDT':
        return amount / usdt_avg
    elif source_currency == 'USDT' and target_currency == 'MYR':
        return amount * myr_avg
    else:
        # 如果有其它币种，则需要扩展逻辑
        return amount

# ================== Excel报表生成工具函数 ==================
def generate_excel_buffer(sheets_data, sheet_order):
    """生成带样式的Excel文件缓冲区"""
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认创建的空工作表

    for sheet_name in sheet_order:
        df = sheets_data[sheet_name]
        ws = wb.create_sheet(sheet_name)

        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # ==== 设置样式 ====
        # 1. 标题行样式（加粗+灰色背景）
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')

        # 2. 遍历每一行，根据内容设置样式
        for row in ws.iter_rows(min_row=2):
            # 判断行类型
            row_type = None
            if row[1].value and '客户支付' in row[1].value:  # B列（订单号）包含"客户支付"
                row_type = 'customer_payment'
            elif row[1].value and '公司支付' in row[1].value:
                row_type = 'company_payment'
            elif row[0].value == '当前余额':  # A列（日期）为"当前余额"
                row_type = 'final_balance'
            elif row[1].value == '期初余额':  # B列（订单号）为"期初余额"
                row_type = 'initial_balance'

            # 应用行背景色
            if row_type == 'initial_balance' or row_type == 'final_balance':
                for cell in row:
                    cell.fill = HIGHLIGHT_FILL
            elif row_type == 'customer_payment':
                for cell in row:
                    cell.fill = CUSTOMER_PAYMENT_FILL
            elif row_type == 'company_payment':
                for cell in row:
                    cell.fill = COMPANY_PAYMENT_FILL

            # 3. 数值列（余额列）设置正负颜色
            for cell in row[8:]:  # 从第I列（索引8）开始是余额列
                if cell.value is None:
                    continue
                try:
                    value = float(cell.value.replace(',', ''))  # 去除千分位逗号
                    if value > 0:
                        cell.font = POSITIVE_FONT
                    elif value < 0:
                        cell.font = NEGATIVE_FONT
                except:
                    pass

        # 设置列宽自适应
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # 保存到缓冲区
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# 通用状态判断函数
def get_tx_status(tx):
    if tx.operator == '/':
        total_quote = tx.amount / tx.rate
    else:
        total_quote = tx.amount * tx.rate

    # 获取结算金额
    settled_base = tx.settled_out if tx.transaction_type == 'buy' else tx.settled_in
    settled_quote = tx.settled_in if tx.transaction_type == 'buy' else tx.settled_out
    
    # 计算整数部分
    base_done = int(settled_base) >= int(tx.amount)
    quote_done = int(settled_quote) >= int(total_quote)
    
    # 计算进度百分比
    base_progress = settled_base / tx.amount if tx.amount != 0 else 0
    quote_progress = settled_quote / total_quote if total_quote != 0 else 0
    min_progress = min(base_progress, quote_progress)
    
    # 状态判断
    if base_done and quote_done:
        return "已完成", min_progress
    elif min_progress > 0:
        return f"部分结算 ({min_progress:.1%})", min_progress
    else:
        return "未结算", min_progress
    
def update_usdt_cost(session, usdt_received: float, myr_spent: float):
    """更新 USDT 的平均成本"""
    record = session.query(USDTAverageCost).first()
    
    if not record:  # 如果没有记录，创建一条新记录
        record = USDTAverageCost(total_usdt=0.0, total_myr_spent=0.0, average_cost=0.0)
        session.add(record)

    # 避免 NoneType 错误
    if record.total_usdt is None:
        record.total_usdt = 0.0
    if record.total_myr_spent is None:
        record.total_myr_spent = 0.0

    record.total_usdt += usdt_received
    record.total_myr_spent += myr_spent

    # 计算平均成本
    if record.total_usdt > 0:
        record.average_cost = record.total_myr_spent / record.total_usdt
    else:
        record.average_cost = 0.0

    session.commit()

def update_myr_cost(session, myr_received: float, usdt_spent: float):
    """更新 MYR 的平均成本"""
    record = session.query(MYRAverageCost).first()
    
    if not record:
        record = MYRAverageCost(total_myr=0.0, total_usdt_spent=0.0, average_cost=0.0)
        session.add(record)

    # 避免 NoneType 错误
    if record.total_myr is None:
        record.total_myr = 0.0
    if record.total_usdt_spent is None:
        record.total_usdt_spent = 0.0

    record.total_myr += myr_received
    record.total_usdt_spent += usdt_spent

    # 计算平均成本
    if record.total_myr > 0:
        record.average_cost = record.total_usdt_spent / record.total_myr
    else:
        record.average_cost = 0.0

    session.commit()

def get_effective_rate(tx):
    """
    计算交易记录的"有效汇率"：
    如果 operator 为 '/'，则有效汇率 = 1/tx.rate；否则直接为 tx.rate。
    """
    try:
        return 1 / float(tx.rate) if tx.operator == '/' else float(tx.rate)
    except Exception as e:
        logger.error(f"计算有效汇率失败: {e}")
        return float(tx.rate)

def generate_detailed_pnl_report_v2(session, start_date, end_date):
    """
    生成详细盈亏报表（修正版本）：
      1. 使用 match_buy_order 对每个买入订单与卖出订单进行匹配
      2. 修正了跨币种转换方向
      3. 每笔订单只使用一次对应的资金

    返回一个列表，每个元素为一行报表数据的字典。
    """
    # 获取实时平均成本（必须存在记录）
    usdt_avg_record = session.query(USDTAverageCost).first()
    myr_avg_record = session.query(MYRAverageCost).first()
    if not usdt_avg_record or not myr_avg_record:
        raise ValueError("平均成本记录缺失，请先完成交易")

    # 汇率定义（例如：1 USDT = 4.42 MYR，1 MYR = 0.226 USDT）
    usdt_avg = usdt_avg_record.average_cost  # 示例值 4.42
    myr_avg = myr_avg_record.average_cost      # 示例值 0.226

    # 构造买入订单列表
    buy_orders = [
        {
            'order_id': tx.order_id,
            'timestamp': tx.timestamp,
            'customer': tx.customer_name,
            'base_currency': tx.base_currency,
            'quote_currency': tx.quote_currency.upper(),
            'amount': float(tx.amount),
            'revenue': float(tx.settled_in),
            'remaining': float(tx.amount)  # 初始化剩余待匹配金额
        }
        for tx in session.query(Transaction).filter(
            Transaction.timestamp.between(start_date, end_date),
            Transaction.transaction_type == 'buy'
        ).order_by(Transaction.timestamp.asc())
    ]

    # 构造卖出订单列表（深拷贝，确保后续匹配中修改不会影响原数据）
    sell_orders = deepcopy([
        {
            'order_id': tx.order_id,
            'base_currency': tx.base_currency,
            'quote_currency': tx.quote_currency.upper(),
            'amount': float(tx.amount),
            'original_cost': float(tx.settled_out),  # 记录原始成本
            'remaining': float(tx.amount)  # 初始化剩余待分配金额
        }
        for tx in session.query(Transaction).filter(
            Transaction.timestamp.between(start_date, end_date),
            Transaction.transaction_type == 'sell'
        ).order_by(Transaction.timestamp.asc())
    ])

    report_rows = []

    # 针对每个买入订单，调用 match_buy_order 从整个卖出订单池中匹配
    for buy in buy_orders:
        # 调用辅助函数进行匹配（注意sell_orders列表中的 remaining 已在上一个买单中更新）
        cost_usdt, cost_myr, matched_sell_ids = match_buy_order(buy, sell_orders, usdt_avg, myr_avg)

        # 根据买单的支付币种确定盈利计算：只计算对应币种的盈利，另一个币种设为0
        if buy['quote_currency'] == 'USDT':
            profit_usdt = buy['revenue'] - cost_usdt
            profit_myr = 0.0
        else:
            profit_myr = buy['revenue'] - cost_myr
            profit_usdt = 0.0

        report_rows.append({
            '日期': buy['timestamp'].strftime('%Y-%m-%d'),
            '客户姓名': buy['customer'],
            '买入订单': buy['order_id'],
            '买入货币': buy['base_currency'],
            '订单金额': f"{buy['amount']:,.2f}",
            '客户支付': f"{buy['revenue']:,.2f} {buy['quote_currency']}",
            'USDT成本': f"{cost_usdt:,.2f}" if buy['quote_currency'] == 'USDT' else "0.00",
            'MYR成本': f"{cost_myr:,.2f}" if buy['quote_currency'] == 'MYR' else "0.00",
            '实际盈利（USDT）': f"{profit_usdt:,.2f}",
            '实际盈利（MYR）': f"{profit_myr:,.2f}",
            '匹配卖出订单': ','.join(matched_sell_ids)
        })

    return report_rows

def generate_detailed_pnl_report_v3(session, start_date, end_date):
    """
    Generate a detailed P&L report that shows:
    1. Each buy transaction
    2. The actual cost basis (in both USDT and MYR)
    3. The matched sell transactions
    4. The realized profit/loss per transaction
    """
    report_rows = []
    
    # Get all buy transactions in the period
    buy_orders = session.query(Transaction).filter(
        Transaction.timestamp.between(start_date, end_date),
        Transaction.transaction_type == 'buy'
    ).order_by(Transaction.timestamp.asc()).all()
    
    # Get ALL sell transactions (not just in period) for matching
    sell_orders = session.query(Transaction).filter(
        Transaction.transaction_type == 'sell'
    ).order_by(Transaction.timestamp.asc()).all()
    
    for buy in buy_orders:
        # Calculate buy details
        buy_base_amount = buy.amount  # e.g., 1000 USDT
        buy_quote_amount = (buy.amount / buy.rate if buy.operator == '/' 
                          else buy.amount * buy.rate)  # e.g., 4420 MYR
        
        # Find all sells that could match this buy
        matched_sells = []
        remaining_amount = buy_base_amount
        realized_quote = 0  # Amount received in quote currency
        
        for sell in sell_orders:
            if sell.base_currency != buy.base_currency:
                continue
                
            # Calculate how much of this sell can match with remaining buy amount
            match_amount = min(remaining_amount, sell.amount)
            if match_amount <= 0:
                continue
                
            # Calculate the sell proceeds for this match
            sell_quote_amount = (match_amount / sell.rate if sell.operator == '/' 
                               else match_amount * sell.rate)
            
            matched_sells.append({
                'order_id': sell.order_id,
                'timestamp': sell.timestamp,
                'amount': match_amount,
                'rate': sell.rate,
                'proceeds': sell_quote_amount
            })
            
            remaining_amount -= match_amount
            realized_quote += sell_quote_amount
            
            if remaining_amount <= 0:
                break
        
        # Calculate P&L
        if buy.quote_currency == 'USDT':
            cost_usdt = buy_quote_amount
            cost_myr = buy_quote_amount * get_usdt_rate(session, buy.timestamp)
        else:  # MYR
            cost_myr = buy_quote_amount
            cost_usdt = buy_quote_amount * get_myr_rate(session, buy.timestamp)
            
        realized_profit_quote = realized_quote - buy_quote_amount
        
        report_rows.append({
            '日期': buy.timestamp.strftime('%Y-%m-%d'),
            '买入订单': buy.order_id,
            '买入数量': f"{buy_base_amount:,.2f} {buy.base_currency}",
            '买入汇率': f"{buy.rate:.4f}",
            '支付金额': f"{buy_quote_amount:,.2f} {buy.quote_currency}",
            'USDT成本': f"{cost_usdt:,.2f}",
            'MYR成本': f"{cost_myr:,.2f}",
            '已匹配数量': f"{(buy_base_amount - remaining_amount):,.2f} {buy.base_currency}",
            '已实现收入': f"{realized_quote:,.2f} {buy.quote_currency}",
            '已实现盈亏': f"{realized_profit_quote:+,.2f} {buy.quote_currency}",
            '匹配卖单': '\n'.join([
                f"{s['order_id']} ({s['amount']:,.2f} @ {s['rate']:.4f})" 
                for s in matched_sells
            ]),
            '匹配率': f"{((buy_base_amount - remaining_amount) / buy_base_amount * 100):.1f}%"
        })
    
    return report_rows

def get_usdt_rate(session, timestamp):
    """Get USDT/MYR rate at given timestamp"""
    record = session.query(USDTAverageCost).filter(
        USDTAverageCost.timestamp <= timestamp
    ).order_by(USDTAverageCost.timestamp.desc()).first()
    return record.average_cost if record else 4.42  # Default rate

def get_myr_rate(session, timestamp):
    """Get MYR/USDT rate at given timestamp"""
    record = session.query(MYRAverageCost).filter(
        MYRAverageCost.timestamp <= timestamp
    ).order_by(MYRAverageCost.timestamp.desc()).first()
    return record.average_cost if record else 0.226  # Default rate

# ================== 交易处理模块 ==================
async def handle_transaction(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """处理交易指令"""
    session = Session()
    try:
        text = update.message.text.strip()
        logger.info(f"收到交易指令: {text}")

        # 修正后的正则表达式
        pattern = (
            r'^(\w+)\s+'  # 客户名
            r'(买|卖|buy|sell)\s+'  # 交易类型
            r'([\d,]+(?:\.\d*)?)([A-Za-z]{3,4})\s*'  # 金额和基础货币（支持小数）
            r'([/*])\s*'  # 运算符
            r'([\d.]+)\s+'  # 汇率
            r'([A-Za-z]{3,4})$'  # 报价货币
        )
        match = re.match(pattern, text, re.IGNORECASE)

        if not match:
            logger.error(f"格式不匹配：{text}")
            await update.message.reply_text(
                "❌ 格式错误！正确示例：\n"
                "`客户A 买 10000USD/4.42 USDT`\n"
                "`客户B 卖 5000EUR*3.45 GBP`\n"
                "`客户C 买 5678MYR/4.42 USDT`（支持无空格）"
            )
            return

        # 解析参数（调整分组索引）
        customer = match.group(1)
        action = match.group(2).lower()
        amount_str = re.sub(r'[^\d.]', '', match.group(3))  # 增强容错处理
        base_currency = match.group(4).upper()
        operator = match.group(5)
        rate_str = match.group(6)
        quote_currency = match.group(7).upper()

        logger.info(f"解析结果: {customer}, {action}, {amount_str}, {base_currency}, {operator}, {rate_str}, {quote_currency}")

        # 类型转换和计算
        try:
            amount = float(amount_str)
            rate = float(rate_str)
            quote_amount = amount / rate if operator == '/' else amount * rate
        except Exception as e:
            await update.message.reply_text(f"❌ 数值错误：{str(e)}")
            return

        # 关键修复：交易方向逻辑
        if action in ('买', 'buy'):
            transaction_type = 'buy'
            # 客户应支付报价货币（USDT），获得基础货币（MYR）
            receive_currency = base_currency   # 客户收到的货币
            pay_currency = quote_currency      # 客户需要支付的货币
        else:
            transaction_type = 'sell'
            # 客户应支付基础货币（MYR），获得报价货币（USDT）
            receive_currency = quote_currency  # 客户收到的货币
            pay_currency = base_currency       # 客户需要支付的货币

        # 创建交易记录
        order_id = generate_order_id(session)
        new_tx = Transaction(
            order_id=order_id,
            customer_name=customer,
            transaction_type=transaction_type,
            base_currency=base_currency,
            quote_currency=quote_currency,
            amount=amount,
            rate=rate,
            status='pending',
            operator=operator,  
            payment_in=0,
            payment_out=0,
            settled_in=0,
            settled_out=0
        )
        session.add(new_tx)

        # 关键修改：更新余额逻辑
        with session.begin_nested():
            session.add(new_tx)
            if transaction_type == 'buy':
                received_curr = quote_currency
                paid_curr = base_currency
                payment_amount = quote_amount
                received_amount = amount
                # 客户获得基础货币（MYR），支付报价货币（USDT）
                update_balance(session, customer, base_currency, amount)
                update_balance(session, customer, quote_currency, -quote_amount)
            else:
                received_curr = base_currency
                paid_curr = quote_currency
                payment_amount = amount
                received_amount = quote_amount
                # 客户支付基础货币（MYR），获得报价货币（USDT）
                update_balance(session, customer, base_currency, -amount)
                update_balance(session, customer, quote_currency, quote_amount)
        
        session.commit()
        # 更新均价逻辑
        base_curr = new_tx.base_currency.upper()
        quote_curr = new_tx.quote_currency.upper()
        currencies = {base_curr, quote_curr}
        if currencies == {'MYR', 'USDT'}:
            # 计算报价金额
            if new_tx.operator == '/':
                quote_amount = new_tx.amount / new_tx.rate
            else:
                quote_amount = new_tx.amount * new_tx.rate

            # 根据交易类型和货币对更新成本
            if base_curr == 'MYR' and quote_curr == 'USDT':
                if new_tx.transaction_type == 'buy':
                    # 公司获得USDT，支出MYR
                    update_usdt_cost(session, quote_amount, new_tx.amount)
                else:
                    # 公司获得MYR，支出USDT
                    update_myr_cost(session, new_tx.amount, quote_amount)
            elif base_curr == 'USDT' and quote_curr == 'MYR':
                if new_tx.transaction_type == 'buy':
                    # 公司获得MYR，支出USDT
                    update_myr_cost(session, quote_amount, new_tx.amount)
                else:
                    # 公司获得USDT，支出MYR
                    update_usdt_cost(session, new_tx.amount, quote_amount)

        # 成功响应（保持原格式）
        await update.message.reply_text(
            f"✅ *交易成功创建* 🎉\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"▪️ 客户：{customer}\n"
            f"▪️ 单号：`{order_id}`\n"
            f"▪️ 类型：{'买入' if transaction_type == 'buy' else '卖出'}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"💱 *汇率说明*\n"
            f"1 {quote_currency} = {rate:.4f} {base_currency}\n\n"
    
            f"📥 *客户需要支付*：\n"
            f"- {payment_amount:,.2f} {pay_currency}\n"
            f"📤 *客户将获得*：\n" 
            f"- {received_amount:,.2f} {receive_currency}\n\n"
    
            f"🏦 *公司账务变动*：\n"
            f"▸ 收入：{payment_amount:,.2f} {pay_currency}\n"
            f"▸ 支出：{received_amount:,.2f} {receive_currency}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"🔧 *后续操作指引*\n"
            f"1️⃣ 当收到客户款项时：\n"
            f"   `/received {customer} {payment_amount:.2f}{pay_currency}`\n\n"
            f"2️⃣ 当向客户支付时：\n"
            f"   `/paid {customer} {received_amount:.2f}{receive_currency}`\n\n"
            f"📝 支持分次操作，金额可修改"
            
        )

    except Exception as e:
        session.rollback()
        logger.error(f"交易处理失败：{str(e)}", exc_info=True)
        await update.message.reply_text(
            "❌ 交易创建失败！\n"
            "⚠️ 错误详情请查看日志"
        )
    finally:
        Session.remove()


# ───────── 收款命令 /received ─────────
async def handle_received(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    处理 /received 命令：客户支付资金给公司。
    调整逻辑：
      ① 先判断是否存在待结算的订单，其报价币与传入币种匹配（对冲+结算买入订单）【分支 A】；
      ② 在【分支 A】处理完后，如果还有剩余，则再处理传入币种作为卖出订单基础币的订单【分支 B】；
      ③ 若仍有剩余，直接记入余额。
    """
    session = Session()
    response_lines = []
    try:
        args = context.args
        if len(args) < 2:
            await update.message.reply_text("❌ 参数错误！格式: /received [客户] [金额+币种]")
            return

        customer, amount_curr = args[0], args[1]
        try:
            input_amount = float(re.sub(r'[^\d.]', '', amount_curr))
            currency = re.search(r'[A-Za-z]{3,4}', amount_curr, re.I).group().upper()
        except (ValueError, AttributeError):
            await update.message.reply_text("❌ 金额格式错误！示例: /received 客户A 1000USD")
            return

        getcontext().prec = 10
        payment = Decimal(str(input_amount)).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
        response_lines.append(f"【客户 {customer} 收款 {payment:,.2f} {currency}】")
        
        # 先处理【分支 A】：当传入币种与订单报价币匹配时
        # 对冲卖出订单（补齐公司应付部分） + 结算买入订单（客户支付部分）
        offset_total = Decimal('0.00')
        # 判断是否有待结算订单，其报价币与传入币种匹配
        quote_order = session.query(Transaction).filter(
            Transaction.customer_name == customer,
            Transaction.quote_currency == currency,
            Transaction.status.in_(['pending', 'partial'])
        ).first()
        remaining_payment = payment  # 初始传入金额
        
        if quote_order:
            # 对冲卖出订单部分（仅针对卖出订单）
            response_lines.append("---------- 对冲卖出订单 ----------")
            sell_offset_txs = session.query(Transaction).filter(
                Transaction.customer_name == customer,
                Transaction.transaction_type == 'sell',
                Transaction.quote_currency == currency,
                Transaction.status.in_(['pending', 'partial'])
            ).order_by(Transaction.timestamp.asc()).with_for_update().all()
            for tx in sell_offset_txs:
                amt = Decimal(str(tx.amount))
                rate = Decimal(str(tx.rate))
                if tx.operator == '/':
                    expected = (amt / rate).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                else:
                    expected = (amt * rate).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                settled = Decimal(str(tx.settled_out or 0))
                remain_order = expected - settled
                if remain_order <= Decimal('0'):
                    continue
                # 对冲操作：补齐公司应付部分
                tx.settled_out = float(expected)
                tx.status = 'settled' if is_fully_settled(tx) else 'partial'
                offset_total += remain_order
                # 客户支付报价币，余额作反向调整（减少负债）
                update_balance(session, customer, currency, -float(remain_order))
                update_balance(session, 'COMPANY', currency, -float(remain_order))
                session.commit()
                response_lines.append(
                    f"订单 {tx.order_id}（卖出）：对冲结算 {remain_order:,.2f} {currency}，累计支付 {expected:,.2f} {currency}，状态：{tx.status}"
                )
            response_lines.append(f"对冲总额：{offset_total:,.2f} {currency}")
            
            effective_payment = payment + offset_total
            response_lines.append("---------- 结算买入订单 ----------")
            response_lines.append(f"传入金额 + 对冲额 = {effective_payment:,.2f} {currency}")
            
            # 结算客户支付部分的买入订单（报价币匹配）
            processed_orders = set()
            temp_payment = effective_payment
            while temp_payment > Decimal('0'):
                tx = session.query(Transaction).filter(
                    Transaction.customer_name == customer,
                    Transaction.transaction_type == 'buy',
                    Transaction.quote_currency == currency,
                    Transaction.status.in_(['pending', 'partial']),
                    ~Transaction.order_id.in_(processed_orders)
                ).order_by(Transaction.timestamp.asc()).with_for_update().first()
                if not tx:
                    break
                amt = Decimal(str(tx.amount))
                rate = Decimal(str(tx.rate))
                if tx.operator == '/':
                    expected = (amt / rate).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                else:
                    expected = (amt * rate).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                settled = Decimal(str(tx.settled_in or 0))
                remain_order = expected - settled
                if remain_order <= Decimal('0'):
                    processed_orders.add(tx.order_id)
                    continue
                settle_amt = min(temp_payment, remain_order)
                new_settled = (settled + settle_amt).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                tx.settled_in = float(new_settled)
                tx.status = 'settled' if is_fully_settled(tx) else 'partial'
                processed_orders.add(tx.order_id)
                # 客户支付部分，余额作反向调整
                update_balance(session, customer, currency, +float(settle_amt))
                update_balance(session, 'COMPANY', currency, +float(settle_amt))
                session.commit()
                response_lines.append(
                    f"订单 {tx.order_id}（买入）：结算 {settle_amt:,.2f} {currency}，累计支付 {new_settled:,.2f}/{expected:,.2f} {currency}，状态：{tx.status}"
                )
                temp_payment = (temp_payment - settle_amt).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
            remaining_payment = temp_payment  # 分支 A结束后的剩余金额
        
        # 【分支 B】——当传入币种与卖出订单的基础币匹配时，处理剩余部分
        buy_sell_orders = session.query(Transaction).filter(
            Transaction.customer_name == customer,
            Transaction.transaction_type == 'sell',
            Transaction.base_currency == currency,
            Transaction.status.in_(['pending', 'partial'])
        ).order_by(Transaction.timestamp.asc()).with_for_update().all()
        if buy_sell_orders and remaining_payment > Decimal('0'):
            response_lines.append("---------- 结算卖出订单（客户支付部分） ----------")
            for tx in buy_sell_orders:
                if remaining_payment <= Decimal('0'):
                    break
                expected = Decimal(str(tx.amount)).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                settled = Decimal(str(tx.settled_in or 0)).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                remain_order = expected - settled
                if remain_order <= Decimal('0'):
                    continue
                settle_amt = min(remaining_payment, remain_order)
                new_settled = (settled + settle_amt).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                tx.settled_in = float(new_settled)
                tx.status = 'settled' if is_fully_settled(tx) else 'partial'
                update_balance(session, customer, currency, +float(settle_amt))
                update_balance(session, 'COMPANY', currency, +float(settle_amt))
                session.commit()
                response_lines.append(
                    f"订单 {tx.order_id}（卖出）：结算 {settle_amt:,.2f} {currency}（客户支付部分），累计支付 {new_settled:,.2f}/{expected:,.2f} {currency}，状态：{tx.status}"
                )
                remaining_payment = (remaining_payment - settle_amt).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
        # 若还有剩余，直接记入余额
        if remaining_payment > Decimal('0'):
            update_balance(session, customer, currency, +float(remaining_payment))
            update_balance(session, 'COMPANY', currency, +float(remaining_payment))
            session.commit()
            response_lines.append(f"剩余 {remaining_payment:,.2f} {currency}直接计入余额。")
        
        if payment > Decimal('0'):
            payment_record = Transaction(
                order_id=generate_payment_id(session, 'PAY-R'),
                customer_name=customer,
                transaction_type='payment',
                sub_type='客户支付',
                base_currency='-',  # 不适用，可置为占位符
                quote_currency=currency,
                amount=float(payment),  # 支付金额
                rate=0,
                operator='-',  # 占位
                status='-',    # 无进度状态
                timestamp=datetime.now(),
                settled_in=float(payment),  # 记录支付的金额
                settled_out=0
            )
            session.add(payment_record)
            session.commit()
            response_lines.append(
                f"生成支付记录：{payment_record.order_id} - 客户支付 {payment:,.2f} {currency}"
            )
        final_response = [f"✅ 成功处理 {customer} 收款 {payment:,.2f} {currency}", "━━━━━━━━━━━━━━━━━━"] + response_lines
        await update.message.reply_text("\n".join(final_response))
    except Exception as e:
        session.rollback()
        logger.error(f"收款处理失败: {str(e)}", exc_info=True)
        await update.message.reply_text("❌ 操作失败，详情请查看日志")
    finally:
        Session.remove()


# ───────── 付款命令 /paid ─────────
async def handle_paid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    处理 /paid 命令：公司向客户支付资金。
    调整逻辑：
      ① 首先检查传入币种是否用于结算卖出订单（报价币匹配）【分支 A】：
            对冲买入订单（补齐客户支付部分）后，按 FIFO 结算卖出订单；
      ② 在【分支 A】处理完后，如果仍有剩余，则查询待结算的买入订单，
          其中传入币种为买入订单的基础币【分支 B】，继续按 FIFO 结算；
      ③ 最后，剩余金额直接从余额中扣除。
    """
    session = Session()
    response_lines = []
    try:
        args = context.args
        if len(args) < 2:
            await update.message.reply_text("❌ 参数错误！格式: /paid [客户] [金额+币种]")
            return

        customer, amount_curr = args[0], args[1]
        try:
            input_amount = float(re.sub(r'[^\d.]', '', amount_curr))
            currency = re.search(r'[A-Za-z]{3,4}', amount_curr, re.I).group().upper()
        except (ValueError, AttributeError):
            await update.message.reply_text("❌ 金额格式错误！示例: /paid 客户A 1000USD")
            return

        getcontext().prec = 10
        payment = Decimal(str(input_amount)).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
        response_lines.append(f"【客户 {customer} 支付指令，传入金额 {payment:,.2f} {currency}】")
        
        total_offset = Decimal('0.00')
        remaining_payment = payment  # 初始金额

        # 【分支 A】——当传入币种用于卖出订单（报价币匹配）时处理
        sell_orders = session.query(Transaction).filter(
            Transaction.customer_name == customer,
            Transaction.transaction_type == 'sell',
            Transaction.quote_currency == currency,
            Transaction.status.in_(['pending', 'partial'])
        ).order_by(Transaction.timestamp.asc()).with_for_update().all()
        
        if sell_orders:
            # 对冲买入订单部分（针对报价币为传入币种的买入订单）
            response_lines.append("---------- 对冲买入订单 ----------")
            buy_offset_txs = session.query(Transaction).filter(
                Transaction.customer_name == customer,
                Transaction.transaction_type == 'buy',
                Transaction.quote_currency == currency,
                Transaction.status.in_(['pending', 'partial'])
            ).order_by(Transaction.timestamp.asc()).with_for_update().all()
            for tx in buy_offset_txs:
                amt = Decimal(str(tx.amount))
                rate = Decimal(str(tx.rate))
                if tx.operator == '/':
                    expected = (amt / rate).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                else:
                    expected = (amt * rate).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                settled = Decimal(str(tx.settled_in or 0))
                remain_order = expected - settled
                if remain_order <= Decimal('0'):
                    continue
                tx.settled_in = float(expected)
                tx.status = 'settled' if is_fully_settled(tx) else 'partial'
                total_offset += remain_order
                update_balance(session, customer, currency, float(remain_order))
                update_balance(session, 'COMPANY', currency, float(remain_order))
                session.commit()
                response_lines.append(
                    f"订单 {tx.order_id}（买入）：对冲结算 {remain_order:,.2f} {currency}，累计结清 {expected:,.2f} {currency}，状态：{tx.status}"
                )
            response_lines.append(f"对冲总额：{total_offset:,.2f} {currency}")
            
            effective_payment = payment + total_offset
            response_lines.append("---------- 结算卖出订单 ----------")
            response_lines.append(f"传入金额 + 对冲额 = {effective_payment:,.2f} {currency}")
            processed_orders = set()
            temp_payment = effective_payment
            for tx in sell_orders:
                if temp_payment <= Decimal('0'):
                    break
                amt = Decimal(str(tx.amount))
                rate = Decimal(str(tx.rate))
                if tx.operator == '/':
                    expected = (amt / rate).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                else:
                    expected = (amt * rate).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                settled = Decimal(str(tx.settled_out or 0))
                remain_order = expected - settled
                if remain_order <= Decimal('0'):
                    processed_orders.add(tx.order_id)
                    continue
                settle_amt = min(temp_payment, remain_order)
                new_settled = (settled + settle_amt).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                tx.settled_out = float(new_settled)
                tx.status = 'settled' if is_fully_settled(tx) else 'partial'
                processed_orders.add(tx.order_id)
                update_balance(session, customer, currency, -float(settle_amt))
                update_balance(session, 'COMPANY', currency, -float(settle_amt))
                session.commit()
                response_lines.append(
                    f"订单 {tx.order_id}（卖出）：结算 {settle_amt:,.2f} {currency}，累计支付 {new_settled:,.2f}/{expected:,.2f} {currency}，状态：{tx.status}"
                )
                temp_payment = (temp_payment - settle_amt).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
            remaining_payment = temp_payment  # 分支 A结束后剩余
            
        # 【分支 B】——处理待结算的买入订单（基础币匹配）
        buy_orders = session.query(Transaction).filter(
            Transaction.customer_name == customer,
            Transaction.transaction_type == 'buy',
            Transaction.base_currency == currency,
            Transaction.status.in_(['pending', 'partial'])
        ).order_by(Transaction.timestamp.asc()).with_for_update().all()
        if buy_orders and remaining_payment > Decimal('0'):
            response_lines.append("---------- 结算买入订单（公司支付部分） ----------")
            for tx in buy_orders:
                if remaining_payment <= Decimal('0'):
                    break
                expected = Decimal(str(tx.amount)).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                settled = Decimal(str(tx.settled_out or 0)).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                remain_order = expected - settled
                if remain_order <= Decimal('0'):
                    continue
                settle_amt = min(remaining_payment, remain_order)
                new_settled = (settled + settle_amt).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                tx.settled_out = float(new_settled)
                tx.status = 'settled' if is_fully_settled(tx) else 'partial'
                update_balance(session, customer, currency, -float(settle_amt))
                update_balance(session, 'COMPANY', currency, -float(settle_amt))
                session.commit()
                response_lines.append(
                    f"订单 {tx.order_id}（买入）：结算 {settle_amt:,.2f} {currency}（公司支付部分），累计支付 {new_settled:,.2f}/{expected:,.2f} {currency}，状态：{tx.status}"
                )
                remaining_payment = (remaining_payment - settle_amt).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
        if remaining_payment > Decimal('0'):
            update_balance(session, customer, currency, -float(remaining_payment))
            update_balance(session, 'COMPANY', currency, -float(remaining_payment))
            session.commit()
            response_lines.append(f"剩余 {remaining_payment:,.2f} {currency}直接从余额中扣除。")
        
        if payment > Decimal('0'):
            payment_record = Transaction(
                order_id=generate_payment_id(session, 'PAY-P'),
                customer_name=customer,
                transaction_type='payment',
                sub_type='公司支付',
                base_currency=currency,  # 这里记录支付币种在基础币列（例如支付 MYR）
                quote_currency='-',      # 不适用
                amount=float(payment),
                rate=0,
                operator='-',
                status='-',
                timestamp=datetime.now(),
                settled_in=0,
                settled_out=float(payment)
            )
            session.add(payment_record)
            session.commit()
            response_lines.append(
                f"生成支付记录：{payment_record.order_id} - 公司支付 {payment:,.2f} {currency}"
            )
        final_response = [f"✅ 成功处理 {customer} 支付 {payment:,.2f} {currency}", "━━━━━━━━━━━━━━━━━━"] + response_lines
        await update.message.reply_text("\n".join(final_response))
    except Exception as e:
        session.rollback()
        logger.error(f"付款处理失败: {str(e)}", exc_info=True)
        await update.message.reply_text("❌ 操作失败，详情请查看日志")
    finally:
        Session.remove()

async def cancel_payment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    撤销支付记录，并逆向更新客户和公司的余额。
    用法示例：/cancel_payment PAY-R-1677481234-1234
    """
    session = Session()
    try:
        if not context.args:
            await update.message.reply_text("❌ 需要支付记录订单号！\n用法: /cancel_payment [支付记录订单号]")
            return

        order_id = context.args[0].upper()  # 支付记录的订单号
        # 仅针对支付记录进行查找
        payment_record = session.query(Transaction).filter_by(order_id=order_id, transaction_type='payment').first()
        if not payment_record:
            await update.message.reply_text("❌ 找不到对应的支付记录")
            return

        # 防止重复撤销（假设status标记为'canceled'后表示已撤销）
        if payment_record.status == 'canceled':
            await update.message.reply_text("❌ 此支付记录已被撤销")
            return

        # 根据支付类型逆向更新余额
        if payment_record.sub_type == '客户支付':
            # 原操作：双方余额增加支付金额（通常是报价币）
            # 撤销操作：双方余额减少支付金额
            update_balance(session, payment_record.customer_name, payment_record.quote_currency, -payment_record.settled_in)
            update_balance(session, 'COMPANY', payment_record.quote_currency, -payment_record.settled_in)
        elif payment_record.sub_type == '公司支付':
            # 原操作：双方余额减少支付金额（通常是基础币）
            # 撤销操作：双方余额增加支付金额
            update_balance(session, payment_record.customer_name, payment_record.base_currency, +payment_record.settled_out)
            update_balance(session, 'COMPANY', payment_record.base_currency, +payment_record.settled_out)
        else:
            await update.message.reply_text("❌ 未知的支付类型，无法撤销")
            return

        # 标记该支付记录为已撤销
        payment_record.status = 'canceled'
        session.commit()

        await update.message.reply_text(f"✅ 支付记录 {order_id} 已成功撤销")
    except Exception as e:
        session.rollback()
        logger.error(f"撤销支付记录失败: {str(e)}", exc_info=True)
        await update.message.reply_text(f"❌ 撤销失败: {str(e)}")
    finally:
        Session.remove()

# ================== 余额管理模块 ==================
async def balance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """查询余额"""
    session = Session()
    try:
        customer = context.args[0] if context.args else 'COMPANY'
        balances = session.query(Balance).filter_by(customer_name=customer).all()
        
        if not balances:
            await update.message.reply_text(f"📭 {customer} 当前没有余额记录")
            return
            
        balance_list = "\n".join([f"▫️ {b.currency}: {b.amount:+,.2f} 💵" for b in balances])
        await update.message.reply_text(
            f"📊 *余额报告* 🏦\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"👤 客户：{customer}\n\n"
            f"💰 当前余额：\n"
            f"{balance_list}",
            parse_mode="Markdown"
        )
    
    except Exception as e:
        logger.error(f"余额查询失败: {str(e)}")
        await update.message.reply_text("❌ 查询失败")
    finally:
        Session.remove()

async def adjust_balance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """手动调整余额"""
    session = Session()
    try:
        args = context.args
        if len(args) < 4:
            await update.message.reply_text("❌ 参数错误！格式: /adjust [客户] [货币] [±金额] [备注]")
            return

        customer, currency, amount_str, *note_parts = args
        note = ' '.join(note_parts)
        
        try:
            amount = float(amount_str)
            currency = currency.upper()
        except ValueError:
            await update.message.reply_text("❌ 金额格式错误")
            return

        # 记录调整
        adj = Adjustment(
            customer_name=customer,
            currency=currency,
            amount=amount,
            note=note
        )
        session.add(adj)
        
        # 更新余额
        update_balance(session, customer, currency, amount)
        session.commit()
        
        await update.message.reply_text(
            f"⚖️ *余额调整完成* ✅\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"👤 客户：{customer}\n"
            f"💱 货币：{currency}\n"
            f"📈 调整量：{amount:+,.2f}\n"
            f"📝 备注：{note}"
        )
    except Exception as e:
        session.rollback()
        logger.error(f"余额调整失败: {str(e)}")
        await update.message.reply_text("❌ 调整失败")
    finally:
        Session.remove()

async def list_debts(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """查询欠款明细（排除公司账户）"""
    session = Session()
    try:
        customer = context.args[0] if context.args else None
        query = session.query(Balance).filter(Balance.customer_name != 'COMPANY')
        if customer:
            query = query.filter_by(customer_name=customer)
        
        balances = query.all()
        debt_report = ["📋 *欠款明细报告* ⚠️", "━━━━━━━━━━━━━━━━━━━━"]
        
        grouped = defaultdict(dict)
        for b in balances:
            grouped[b.customer_name][b.currency] = b.amount
        
        for cust, currencies in grouped.items():
            debt_report.append(f"👤 客户: {cust}")
            for curr, amt in currencies.items():
                if amt > 0.01:  # 余额为正 → 公司欠客户
                    debt_report.append(f"▫️ 公司欠客户 {amt:,.2f} {curr} 🟢")
                elif amt < -0.01:  # 余额为负 → 客户欠公司
                    debt_report.append(f"▫️ 客户欠公司 {-amt:,.2f} {curr} 🔴")
            debt_report.append("━━━━━━━━━━━━━━━━━━━━")
        
        await update.message.reply_text("\n".join(debt_report))
    except Exception as e:
        logger.error(f"欠款查询失败: {str(e)}")
        await update.message.reply_text("❌ 查询失败")
    finally:
        Session.remove()
                
# ================== 支出管理模块 ==================
async def add_expense(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """记录公司支出（仅记录，不更新余额）"""
    session = Session()
    try:
        args = context.args
        if len(args) < 2:
            await update.message.reply_text("❌ 参数错误！格式: /expense [金额+货币] [用途]")
            return

        amount_curr, *purpose_parts = args
        purpose = ' '.join(purpose_parts)
        
        try:
            amount = float(re.sub(r'[^\d.]', '', amount_curr))
            currency = re.search(r'[A-Z]{3,4}', amount_curr, re.I).group().upper()
        except (ValueError, AttributeError):
            await update.message.reply_text("❌ 金额格式错误！示例: /expense 100USD 办公室租金")
            return

        expense = Expense(
            amount=amount,
            currency=currency,
            purpose=purpose
        )
        session.add(expense)
        session.commit()
        
        await update.message.reply_text(
            f"💸 *支出记录已添加* ✅\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"💰 金额：{amount:,.2f} {currency}\n"
            f"📝 用途：{purpose}"
        )
    except Exception as e:
        session.rollback()
        logger.error(f"支出记录失败: {str(e)}")
        await update.message.reply_text("❌ 记录失败")
    finally:
        Session.remove()

async def cancel_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """撤销交易并恢复初始余额"""
    session = Session()
    try:
        if not context.args:
            await update.message.reply_text("❌ 需要订单号！用法: /cancel YS000000001")
            return

        order_id = context.args[0].upper()
        tx = session.query(Transaction).filter_by(order_id=order_id).first()
        if not tx:
            await update.message.reply_text("❌ 找不到该交易")
            return

        # 计算实际交易金额（根据运算符）
        if tx.operator == '/':
            quote_amount = tx.amount / tx.rate
        else:
            quote_amount = tx.amount * tx.rate

        # 撤销初始交易影响
        if tx.transaction_type == 'buy':
            # 反向操作：
            update_balance(session, tx.customer_name, tx.base_currency, -tx.amount)  # 扣除获得的基础货币
            update_balance(session, tx.customer_name, tx.quote_currency, quote_amount)  # 恢复支付的报价货币
        else:
            update_balance(session, tx.customer_name, tx.base_currency, tx.amount)  # 恢复支付的基础货币
            update_balance(session, tx.customer_name, tx.quote_currency, -quote_amount)  # 扣除获得的报价货币

        session.delete(tx)
        session.commit()

        await update.message.reply_text(
            f"✅ 交易 {order_id} 已撤销\n"
            f"━━━━━━━━━━━━━━\n"
            f"▸ {tx.base_currency} 调整：{-tx.amount if tx.transaction_type == 'buy' else tx.amount:+,.2f}\n"
            f"▸ {tx.quote_currency} 调整：{quote_amount if tx.transaction_type == 'buy' else -quote_amount:+,.2f}"
        )

    except Exception as e:
        session.rollback()
        logger.error(f"撤销失败: {str(e)}")
        await update.message.reply_text(f"❌ 撤销失败: {str(e)}")
    finally:
        Session.remove()

async def delete_customer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """删除客户及其所有相关数据"""
    session = Session()
    try:
        args = context.args
        if not args:
            await update.message.reply_text("❌ 请输入客户名称，格式: /delete_customer [客户名]")
            return
        customer_name = args[0]

        # 删除所有相关记录（使用事务保证原子性）
        with session.begin_nested():
            # 删除客户基本信息（如果存在）
            customer = session.query(Customer).filter_by(name=customer_name).first()
            if customer:
                session.delete(customer)
                
            # 删除余额记录
            balance_count = session.query(Balance).filter_by(customer_name=customer_name).delete()
            
            # 删除交易记录
            tx_count = session.query(Transaction).filter_by(customer_name=customer_name).delete()
            
            # 删除调整记录
            adj_count = session.query(Adjustment).filter_by(customer_name=customer_name).delete()

        session.commit()

        response = (
            f"✅ 客户 *{customer_name}* 数据已清除\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"▫️ 删除余额记录：{balance_count} 条\n"
            f"▫️ 删除交易记录：{tx_count} 条\n"
            f"▫️ 删除调整记录：{adj_count} 条\n"
            f"▫️ 删除客户资料：{1 if customer else 0} 条\n\n"
            f"⚠️ 该操作不可逆，所有相关数据已从数据库中清除"
        )
        await update.message.reply_text(response, parse_mode="Markdown")

    except Exception as e:
        session.rollback()
        logger.error(f"删除客户失败: {str(e)}", exc_info=True)
        await update.message.reply_text(
            "❌ 删除操作失败！\n"
            "⚠️ 错误详情请查看服务器日志"
        )
    finally:
        Session.remove()

# ================== 支出管理模块（续） ==================
async def list_expenses(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """查询支出记录，可按日期范围查询；若无参数，则默认查询当月记录"""
    session = Session()
    try:
        args = context.args
        from datetime import datetime, timedelta

        if len(args) == 2:
            try:
                start_date = datetime.strptime(args[0], '%Y-%m-%d')
                end_date = datetime.strptime(args[1], '%Y-%m-%d')
                # 增加一天，使 end_date 成为不包含当天的截止日期
                end_date = end_date + timedelta(days=1)
            except Exception:
                await update.message.reply_text("❌ 日期格式错误！请使用 YYYY-MM-DD 格式")
                return
        elif len(args) == 0:
            now = datetime.utcnow()
            start_date = datetime(now.year, now.month, 1)
            # 计算下个月第一天作为截止时间
            if now.month == 12:
                end_date = datetime(now.year + 1, 1, 1)
            else:
                end_date = datetime(now.year, now.month + 1, 1)
        else:
            await update.message.reply_text("❌ 参数错误！格式: /list_expenses [start_date] [end_date] (日期格式: YYYY-MM-DD)")
            return

        expenses = session.query(Expense).filter(
            Expense.timestamp >= start_date,
            Expense.timestamp < end_date
        ).order_by(Expense.timestamp.desc()).all()

        if not expenses:
            await update.message.reply_text("📝 当前无支出记录")
            return

        # 显示标题中 end_date 减一天以展示正确的结束日期
        report = [
            f"📋 支出记录 ({start_date.strftime('%Y-%m-%d')} 至 {(end_date - timedelta(days=1)).strftime('%Y-%m-%d')})",
            "━━━━━━━━━━━━━━━"
        ]
        for exp in expenses:
            report.append(
                f"▫️ {exp.timestamp.strftime('%Y-%m-%d %H:%M')}\n"
                f"金额: {exp.amount:,.2f} {exp.currency}\n"
                f"用途: {exp.purpose}\n"
                "━━━━━━━━━━━━━━━"
            )
        
        full_report = "\n".join(report)
        # 分页发送，防止消息过长
        for i in range(0, len(full_report), 4000):
            await update.message.reply_text(full_report[i:i+4000])
    except Exception as e:
        logger.error(f"支出查询失败: {str(e)}")
        await update.message.reply_text("❌ 查询失败")
    finally:
        Session.remove()

def generate_expenses_image_side_summary(expenses_data, start_date, end_date, total_per_currency):
    """
    生成“支出报表”图片：
      - 左侧：支出明细 (日期、时间、金额、币种、用途)
      - 右侧：各币种支出合计
      - 字体较大，行高加大
    """
    import io
    from PIL import Image, ImageDraw, ImageFont

    # ========== 1) 样式与字体 ==========
    padding = 40
    row_height = 65
    header_height = 140
    border_color = '#000000'
    line_width = 2

    # 颜色
    header_fill = '#E8E8E8'   # 表头
    white_fill = '#FFFFFF'    # 普通行
    summary_fill = '#FFFACD'  # 右侧面板标题行（淡黄）

    try:
        title_font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 40)
        header_font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 28)
        font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 24)
        bold_font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 24, index=1)
    except:
        # 回退
        title_font = ImageFont.load_default()
        header_font = title_font
        font = title_font
        bold_font = title_font

    # ========== 2) 定义列、初始列宽、对齐方式 ==========
    columns = [
        {"key": "日期",  "title": "日期",   "width": 140, "align": "center"},
        {"key": "时间",  "title": "时间",   "width": 100, "align": "center"},
        {"key": "金额",  "title": "金额",   "width": 140, "align": "right"},
        {"key": "币种",  "title": "币种",   "width": 90,  "align": "right"},
        {"key": "用途",  "title": "用途",   "width": 300, "align": "left"},
    ]

    dummy_img = Image.new('RGB', (10, 10), 'white')
    dummy_draw = ImageDraw.Draw(dummy_img)

    def measure_text(txt, fnt):
        bbox = dummy_draw.textbbox((0, 0), txt, font=fnt)
        return (bbox[2] - bbox[0], bbox[3] - bbox[1])

    # ========== 3) 动态调整列宽 ==========
    for col in columns:
        w, h = measure_text(col["title"], header_font)
        col["width"] = max(col["width"], w + 30)

    for row in expenses_data:
        for col in columns:
            cell_text = row[col["key"]]
            w, h = measure_text(cell_text, font)
            if w + 30 > col["width"]:
                col["width"] = w + 30

    table_width = sum(c["width"] for c in columns)
    data_rows = len(expenses_data)

    # ========== 4) 右侧汇总面板 ==========
    # 显示“总支出 (Total Expenses)” + N 行（每种币种）
    summary_lines = 1 + len(total_per_currency)
    summary_width = 340  # 右侧面板固定宽度，可调
    summary_height = summary_lines * row_height

    # ========== 5) 计算画布整体宽高 ==========
    img_width = table_width + summary_width + padding*3
    table_height = header_height + (1 + data_rows)*row_height
    content_height = max(table_height, header_height + summary_height)
    img_height = content_height + padding

    # ========== 6) 创建画布并绘制标题 ==========
    img = Image.new('RGB', (img_width, img_height), color='white')
    draw = ImageDraw.Draw(img)

    title_text = "支出报表"
    date_range_text = f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"
    draw.text((padding, padding), title_text, font=title_font, fill='black')
    draw.text((padding, padding + 60), f"日期范围: {date_range_text}", font=header_font, fill='black')

    table_x = padding
    table_y = padding + header_height

    # ========== 7) 绘制表头 ==========
    x = table_x
    y = table_y
    for col in columns:
        w = col["width"]
        draw.rectangle([x, y, x + w, y + row_height],
                       fill=header_fill, outline=border_color, width=line_width)
        txt_w, txt_h = measure_text(col["title"], header_font)
        if col["align"] == "center":
            text_x = x + (w - txt_w)//2
        elif col["align"] == "right":
            text_x = x + w - txt_w - 10
        else:
            text_x = x + 10
        text_y = y + (row_height - txt_h)//2
        draw.text((text_x, text_y), col["title"], font=header_font, fill='black')
        x += w
    y += row_height

    # ========== 8) 绘制数据行 ==========
    for row_data in expenses_data:
        x = table_x
        for col in columns:
            w = col["width"]
            cell_text = row_data[col["key"]]
            draw.rectangle([x, y, x + w, y + row_height],
                           fill=white_fill, outline=border_color, width=line_width)
            txt_w, txt_h = measure_text(cell_text, font)
            if col["align"] == "center":
                text_x = x + (w - txt_w)//2
            elif col["align"] == "right":
                text_x = x + w - txt_w - 10
            else:
                text_x = x + 10
            text_y = y + (row_height - txt_h)//2
            draw.text((text_x, text_y), cell_text, font=font, fill='black')
            x += w
        y += row_height

    # ========== 9) 右侧汇总面板 ==========
    summary_x = table_x + table_width + padding
    summary_y = padding + header_height

    def draw_summary_line(text, y_offset, fill_color=white_fill, fnt=font):
        draw.rectangle([summary_x, y_offset, summary_x + summary_width, y_offset + row_height],
                       fill=fill_color, outline=border_color, width=line_width)
        txt_w, txt_h = measure_text(text, fnt)
        # 左对齐
        draw.text((summary_x + 10, y_offset + (row_height - txt_h)//2),
                  text, font=fnt, fill='black')

    current_y = summary_y
    # 标题
    draw_summary_line("总支出 (Total Expenses)", current_y, fill_color=summary_fill, fnt=bold_font)
    current_y += row_height

    # 各币种
    for ccy in sorted(total_per_currency.keys()):
        amt = total_per_currency[ccy]
        line_str = f"{ccy}: {amt:,.2f}"
        draw_summary_line(line_str, current_y, fill_color=white_fill, fnt=font)
        current_y += row_height

    # ========== 10) 输出到 BytesIO ==========
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    return buffer

async def list_expenses_image(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    生成支出记录的图片报表。
    用法：
      /expensesimg               -> 默认查询当月
      /expensesimg 2025-03-01 2025-03-15  -> 指定日期范围 (YYYY-MM-DD YYYY-MM-DD)
    """
    session = Session()
    try:
        from datetime import datetime, timedelta

        args = context.args
        if len(args) == 2:
            try:
                start_date = datetime.strptime(args[0], '%Y-%m-%d')
                end_date = datetime.strptime(args[1], '%Y-%m-%d')
                # 为避免漏掉当日 23:59:59，可以把 end_date 调整到这天末
                end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            except Exception:
                await update.message.reply_text("❌ 日期格式错误！请使用 YYYY-MM-DD 格式")
                return
        elif len(args) == 0:
            # 默认查询当月
            now = datetime.now()
            start_date = datetime(now.year, now.month, 1)
            # 下个月的1号减1秒
            if now.month == 12:
                next_month_start = datetime(now.year + 1, 1, 1)
            else:
                next_month_start = datetime(now.year, now.month + 1, 1)
            end_date = next_month_start - timedelta(seconds=1)
        else:
            await update.message.reply_text("❌ 参数错误！格式: /expensesimg [start_date] [end_date]")
            return

        # 查询
        expenses = session.query(Expense).filter(
            Expense.timestamp >= start_date,
            Expense.timestamp <= end_date
        ).order_by(Expense.timestamp.asc()).all()

        if not expenses:
            await update.message.reply_text("📝 指定时间段内无支出记录")
            return

        # 整理数据
        expenses_data = []
        total_per_currency = defaultdict(float)

        for exp in expenses:
            date_str = exp.timestamp.strftime('%Y-%m-%d')
            time_str = exp.timestamp.strftime('%H:%M')
            amount_str = f"{exp.amount:,.2f}"
            currency = exp.currency
            purpose = exp.purpose or ""

            expenses_data.append({
                "日期": date_str,
                "时间": time_str,
                "金额": amount_str,
                "币种": currency,
                "用途": purpose
            })

            # 汇总
            total_per_currency[currency] += exp.amount

        # 调用上面定义的函数生成图片
        img_buffer = generate_expenses_image_side_summary(expenses_data, start_date, end_date, total_per_currency)

        # 发送图片
        await update.message.reply_photo(
            photo=img_buffer,
            caption="📊 支出记录报表"
        )

    except Exception as e:
        session.rollback()
        logger.error(f"支出报表生成失败: {str(e)}", exc_info=True)
        await update.message.reply_text("❌ 报表生成失败，请查看日志")
    finally:
        Session.remove()

# ================== 报表生成模块 ==================
async def pnl_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """生成精准的货币独立盈亏报告（针对订单计算盈亏）"""
    session = Session()
    try:
        # 解析参数
        args = context.args or []
        excel_mode = 'excel' in args
        date_args = [a for a in args if a != 'excel']
        
        # 解析日期范围
        if date_args:
            try:
                start_date, end_date = parse_date_range(' '.join(date_args))
            except ValueError as e:
                await update.message.reply_text(f"❌ {str(e)}")
                return
        else:
            now = datetime.now()
            start_date = now.replace(day=1, hour=0, minute=0, second=0)
            end_date = now.replace(day=calendar.monthrange(now.year, now.month)[1], 
                                hour=23, minute=59, second=59)

        # 获取交易记录和支出记录
        txs = session.query(Transaction).filter(
            Transaction.timestamp.between(start_date, end_date),
            Transaction.transaction_type.in_(["buy", "sell"])
        ).all()
        
        expenses = session.query(Expense).filter(
            Expense.timestamp.between(start_date, end_date)
        ).all()

        # 初始化货币报告
        currency_report = defaultdict(lambda: {
            'actual_income': 0.0,  # 实际收入（已结算）
            'actual_expense': 0.0,  # 实际支出（已结算）
            'pending_income': 0.0,  # 应收未收
            'pending_expense': 0.0,  # 应付未付
            'credit_balance': 0.0,  # 客户多付的信用余额
            'total_income': 0.0,    # 总应收款
            'total_expense': 0.0,   # 总应付款
            'expense': 0.0          # 支出
        })

        # 处理交易记录
        for tx in txs:
            # 根据运算符计算报价货币金额
            if tx.operator == '/':
                total_quote = tx.amount / tx.rate
            else:
                total_quote = tx.amount * tx.rate

            if tx.transaction_type == 'buy':
                # 买入交易：客户支付报价货币，获得基础货币
                currency_report[tx.quote_currency]['total_income'] += total_quote  # 总应收款
                currency_report[tx.quote_currency]['actual_income'] += tx.settled_in  # 已收款
                currency_report[tx.quote_currency]['pending_income'] += total_quote - tx.settled_in  # 应收未收
                currency_report[tx.base_currency]['total_expense'] += tx.amount  # 总应付款
                currency_report[tx.base_currency]['actual_expense'] += tx.settled_out  # 已付款
                currency_report[tx.base_currency]['pending_expense'] += tx.amount - tx.settled_out  # 应付未付
            else:
                # 卖出交易：客户支付基础货币，获得报价货币
                currency_report[tx.base_currency]['total_income'] += tx.amount  # 总应收款
                currency_report[tx.base_currency]['actual_income'] += tx.settled_in  # 已收款
                currency_report[tx.base_currency]['pending_income'] += tx.amount - tx.settled_in  # 应收未收
                currency_report[tx.quote_currency]['total_expense'] += total_quote  # 总应付款
                currency_report[tx.quote_currency]['actual_expense'] += tx.settled_out  # 已付款
                currency_report[tx.quote_currency]['pending_expense'] += total_quote - tx.settled_out  # 应付未付

        # 处理支出记录
        for exp in expenses:
            currency_report[exp.currency]['expense'] += exp.amount
            currency_report[exp.currency]['actual_expense'] += exp.amount

        # 计算客户多付的信用余额
        for currency, data in currency_report.items():
            # 信用余额 = 已收款 - 总应收款
            data['credit_balance'] = max(0, data['actual_income'] - data['total_income'])

        # ================== Excel报表生成 ==================
        if excel_mode:
            # 交易明细
            tx_data = []
            for tx in txs:
                if tx.operator == '/':
                    total_quote = tx.amount / tx.rate
                else:
                    total_quote = tx.amount * tx.rate

                # 结算金额计算
                settled_base = tx.settled_out if tx.transaction_type == 'buy' else tx.settled_in
                settled_quote = tx.settled_in if tx.transaction_type == 'buy' else tx.settled_out
    
                # 计算双货币进度
                base_progress = settled_base / tx.amount if tx.amount != 0 else 0
                quote_progress = settled_quote / total_quote if total_quote != 0 else 0
                min_progress = min(base_progress, quote_progress)
    
                # 状态判断（取整后判断）
                base_done = int(settled_base) >= int(tx.amount)
                quote_done = int(settled_quote) >= int(total_quote)
                status = "已完成" if base_done and quote_done else "进行中"

                tx_data.append({
                    "日期": tx.timestamp.strftime('%Y-%m-%d'),
                    "订单号": tx.order_id,
                    "客户名称": tx.customer_name,
                    "交易类型": '买入' if tx.transaction_type == 'buy' else '卖出',
                    "基础货币总额": f"{tx.amount:,.2f} {tx.base_currency}",
                    "报价货币总额": f"{total_quote:,.2f} {tx.quote_currency}",
                    "已结基础货币": f"{settled_base:,.2f} {tx.base_currency}",
                    "已结报价货币": f"{settled_quote:,.2f} {tx.quote_currency}",  # 新增结算金额
                    "基础货币进度": f"{base_progress:.1%}",
                    "报价货币进度": f"{quote_progress:.1%}",
                    "状态": status
                })

            # 货币汇总
            currency_data = []
            for curr, data in currency_report.items():
                currency_data.append({
                    "货币": curr,
                    "实际收入": f"{data['actual_income']:,.2f}",
                    "实际支出": f"{data['actual_expense']:,.2f}",
                    "应收未收": f"{data['pending_income']:,.2f}",
                    "应付未付": f"{data['pending_expense']:,.2f}",
                    "信用余额": f"{data['credit_balance']:,.2f}",
                    "净盈亏": f"{data['actual_income'] - data['actual_expense']:,.2f}"
                })

            # 支出记录
            expense_data = [{
                "日期": exp.timestamp.strftime('%Y-%m-%d'),
                "金额": f"{exp.amount:,.2f}",
                "货币": exp.currency,
                "用途": exp.purpose
            } for exp in expenses]

            # 生成Excel
            df_dict = {
                "交易明细": pd.DataFrame(tx_data),
                "货币汇总": pd.DataFrame(currency_data),
                "支出记录": pd.DataFrame(expense_data)
            }
            
            excel_buffer = generate_excel_buffer(df_dict, ["交易明细", "货币汇总", "支出记录"])
            await update.message.reply_document(
                document=excel_buffer,
                filename=f"盈亏报告_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx",
                caption="📊 包含货币独立盈亏的Excel报告"
            )
            return
        
        # ================== 生成文本报告 ==================
        report = [
            f"📊 *盈亏报告* ({start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')})",
            f"▫️ 有效交易：{len(txs)}笔 | 支出记录：{len(expenses)}笔",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━"
        ]
        
        for curr, data in currency_report.items():
            profit = data['actual_income'] - data['actual_expense']
            report.append(
                f"🔘 *{curr}* 货币\n"
                f"▸ 实际收入：{data['actual_income']:+,.2f}\n"
                f"▸ 实际支出：{data['actual_expense']:+,.2f}\n"
                f"▸ 应收未收：{data['pending_income']:,.2f}\n"
                f"▸ 应付未付：{data['pending_expense']:,.2f}\n"
                f"▸ 信用余额：{data['credit_balance']:,.2f}\n"
                f"🏁 净盈亏：{profit:+,.2f}\n"
                "━━━━━━━━━━━━━━━━━━"
            )
            
        await update.message.reply_text("\n".join(report))

    except Exception as e:
        logger.error(f"盈亏报告生成失败: {str(e)}", exc_info=True)
        await update.message.reply_text("❌ 报告生成失败，请检查日志")
    finally:
        Session.remove()

async def detailed_pnl_report_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    session = Session()
    try:
        # 解析日期范围（格式要求："DD/MM/YYYY-DD/MM/YYYY"），若无参数则默认当前月份
        args = context.args or []
        if args:
            try:
                start_date, end_date = parse_date_range(' '.join(args))
            except ValueError:
                await update.message.reply_text("❌ 日期格式错误，请使用 DD/MM/YYYY-DD/MM/YYYY 格式")
                return
        else:
            now = datetime.now()
            start_date = now.replace(day=1, hour=0, minute=0, second=0)
            end_date = now.replace(day=calendar.monthrange(now.year, now.month)[1], hour=23, minute=59, second=59)
        
        report_rows = generate_detailed_pnl_report_v2(session, start_date, end_date)
        if not report_rows:
            await update.message.reply_text("⚠️ 指定时间段内无相关交易数据")
            return
        
        # 转换为 DataFrame，并按照要求设置列顺序
        df = pd.DataFrame(report_rows, columns=[
            '日期','客户姓名','买入订单','买入货币','订单金额','客户支付',
            'USDT成本','MYR成本','实际盈利（USDT）','实际盈利（MYR）','匹配卖出订单'
        ])
        
        excel_buffer = generate_excel_buffer({'详细盈亏报表': df}, ['详细盈亏报表'])
        await update.message.reply_document(
            document=excel_buffer,
            filename=f"详细盈亏报表_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx",
            caption="📊 详细盈亏报表\n显示客户支付金额、USDT/MYR成本及实际盈利"
        )
    except Exception as e:
        logger.error(f"详细盈亏报表生成失败: {str(e)}", exc_info=True)
        await update.message.reply_text("❌ 详细盈亏报表生成失败，请检查日志")
    finally:
        Session.remove()

async def generate_detailed_report(update: Update, context: ContextTypes.DEFAULT_TYPE, period: str):
    session = Session()
    try:
        args = context.args or []
        excel_mode = 'excel' in args
        date_args = [a for a in args if a != 'excel']
        
        # 解析日期范围（增强容错）
        if date_args:
            try:
                if '-' in ' '.join(date_args):
                    start_date, end_date = parse_date_range(' '.join(date_args))
                else:
                    single_date = datetime.strptime(' '.join(date_args), '%d/%m/%Y')
                    start_date = single_date.replace(hour=0, minute=0, second=0)
                    end_date = single_date.replace(hour=23, minute=59, second=59)
            except ValueError as e:
                await update.message.reply_text(f"❌ 日期格式错误，请使用 DD/MM/YYYY 或 DD/MM/YYYY-DD/MM/YYYY")
                return
        else:
            now = datetime.now()
            start_date = now.replace(day=1, hour=0, minute=0, second=0)
            end_date = now.replace(day=calendar.monthrange(now.year, now.month)[1], 
                                hour=23, minute=59, second=59)

        # 获取交易记录和客户信用余额
        txs = session.query(Transaction).filter(
            Transaction.timestamp.between(start_date, end_date),
            Transaction.transaction_type.in_(["buy", "sell"])
        ).all()
      
        # 获取所有客户的信用余额
        credit_balances = session.query(
            Balance.customer_name,
            Balance.currency,
            func.sum(Balance.amount).label('credit')
        ).filter(Balance.amount > 0).group_by(Balance.customer_name, Balance.currency).all()

        # Excel生成修正
        if excel_mode:
            tx_data = []
            for tx in txs:
                try:
                    # 计算应付总额和信用余额
                    if tx.operator == '/':
                        total_quote = tx.amount / tx.rate
                    else:
                        total_quote = tx.amount * tx.rate
                        
                    # 获取该客户的信用余额
                    credit = next(
                        (cb.credit for cb in credit_balances 
                         if cb.customer_name == tx.customer_name 
                         and cb.currency == tx.quote_currency),
                        0.0
                    )

                    # 根据交易类型确定结算逻辑
                    if tx.transaction_type == 'buy':
                        # 买入交易：客户应支付报价货币
                        required = total_quote
                        settled = tx.settled_in
                        credit_used = min(credit, required - settled)
                    else:
                        # 卖出交易：客户应支付基础货币
                        required = tx.amount
                        settled = tx.settled_in
                        credit_used = min(credit, required - settled)

                    # 计算实际需要支付的金额
                    actual_payment = settled + credit_used
                    remaining = required - actual_payment
                    progress = actual_payment / required if required != 0 else 0

                    # 判断状态
                    if tx.transaction_type == 'buy':
                        # 买入交易判断逻辑
                        base_done = int(tx.settled_out) >= int(tx.amount)  # 公司支付的基础货币
                        quote_done = int(tx.settled_in) >= int(total_quote)  # 客户支付的报价货币
                    else:
                        # 卖出交易判断逻辑
                        base_done = int(tx.settled_in) >= int(tx.amount)    # 客户支付的基础货币
                        quote_done = int(tx.settled_out) >= int(total_quote) # 公司支付的报价货币

                    status = "已完成" if base_done and quote_done else "进行中"    

                    if tx.transaction_type == 'buy':
                        settled_base = tx.settled_out  # 公司已支付的基础货币
                        settled_quote = tx.settled_in  # 客户已支付的报价货币
                    else:
                        settled_base = tx.settled_in   # 客户已支付的基础货币
                        settled_quote = tx.settled_out # 公司已支付的报价货币                        

                    # 计算汇率
                    if tx.operator == '/':
                        exchange_rate = tx.rate
                    else:
                        exchange_rate = 1 / tx.rate

                    record = {
                        "日期": tx.timestamp.strftime('%Y-%m-%d'),
                        "订单号": tx.order_id,
                        "客户名称": tx.customer_name,
                        "交易类型": '买入' if tx.transaction_type == 'buy' else '卖出',
                        "基础货币总额": f"{tx.amount:,.2f} {tx.base_currency}",
                        "报价货币总额": f"{total_quote:,.2f} {tx.quote_currency}",
                        "已结基础货币": f"{settled_base:,.2f} {tx.base_currency}",
                        "已结报价货币": f"{settled_quote:,.2f} {tx.quote_currency}", 
                        "基础货币进度": f"{(tx.settled_out / tx.amount * 100):.1f}%" if tx.transaction_type == 'buy' else f"{(tx.settled_in / tx.amount * 100):.1f}%",
                        "报价货币进度": f"{(tx.settled_in / total_quote * 100):.1f}%" if tx.transaction_type == 'buy' else f"{(tx.settled_out / total_quote * 100):.1f}%",
                        "汇率": f"{exchange_rate:.6f}",  # 添加汇率信息
                        "状态": status  # 使用新的状态判断
                    }
                    tx_data.append(record)
                except Exception as e:
                    logger.error(f"处理交易 {tx.order_id} 失败: {str(e)}")
                    continue
            
            if not tx_data:
                await update.message.reply_text("⚠️ 该时间段内无交易记录")
                return

            # 生成信用余额表
            credit_data = [{
                "客户名称": cb.customer_name,
                "货币": cb.currency,
                "信用余额": f"{cb.credit:,.2f}"
            } for cb in credit_balances]

            df_dict = {
                "交易明细": pd.DataFrame(tx_data),
                "信用余额": pd.DataFrame(credit_data)
            }
            
            excel_buffer = generate_excel_buffer(df_dict, ["交易明细", "信用余额"])
            await update.message.reply_document(
                document=excel_buffer,
                filename=f"交易明细_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx",
                caption="📊 包含信用对冲的Excel交易明细"
            )
            return

        # 文本报告生成
        report = [
            f"📋 交易结算明细报告 ({start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
            f"总交易数: {len(txs)}",
            "━━━━━━━━━━━━━━━━━━"
        ]
        
        for tx in txs:
            # 计算应付总额
            if tx.operator == '/':
                total_quote = tx.amount / tx.rate
            else:
                total_quote = tx.amount * tx.rate

            if tx.transaction_type == 'buy':
                settled_base = tx.settled_out  # 公司已支付的基础货币
                settled_quote = tx.settled_in  # 客户已支付的报价货币
            else:
                settled_base = tx.settled_in   # 客户已支付的基础货币
                settled_quote = tx.settled_out # 公司已支付的报价货币

            # 计算进度
            base_progress = settled_base / tx.amount if tx.amount != 0 else 0
            quote_progress = settled_quote / total_quote if total_quote != 0 else 0

            # 判断状态
            base_done = int(settled_base) >= int(tx.amount)
            quote_done = int(settled_quote) >= int(total_quote)
            status = "✅ 已完成" if base_done and quote_done else "🟡 进行中"

            # 添加到报告
            report.append(
                f"📌 {tx.timestamp.strftime('%d/%m %H:%M')} {tx.order_id}\n"
                f"{tx.customer_name} {'买入' if tx.transaction_type == 'buy' else '卖出'} "
                f"{tx.amount:,.2f} {tx.base_currency} @ {tx.rate:.4f}\n"
                f"├─ 已结基础货币: {settled_base:,.2f}/{tx.amount:,.2f} {tx.base_currency} ({base_progress:.1%})\n"
                f"├─ 已结报价货币: {settled_quote:,.2f}/{total_quote:,.2f} {tx.quote_currency} ({quote_progress:.1%})\n"
                f"└─ 状态: {status}"
                "━━━━━━━━━━━━━━━━━━"
            )
        
        # 发送报告
        full_report = "\n".join(report)
        for i in range(0, len(full_report), 4000):
            await update.message.reply_text(full_report[i:i+4000])
    except Exception as e:
        logger.error(f"交易报表生成失败: {str(e)}")
        await update.message.reply_text("❌ 生成失败")
    finally:
        Session.remove()
                      
async def customer_statement(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Generate customer statement as Excel or Image"""
    session = Session()
    try:
        args = context.args or []
        if not args:
            await update.message.reply_text("❌ 需要客户名称！格式: /creport [客户名] [日期范围] [excel/image]")
            return

        excel_mode = 'excel' in args
        image_mode = 'image' in args
        clean_args = [a for a in args if a not in ['excel', 'image']]
        customer = clean_args[0]
        date_args = clean_args[1:]

        # 解析日期范围
        if date_args:
            try:
                start_date, end_date = parse_date_range(' '.join(date_args))
            except ValueError as e:
                await update.message.reply_text(f"❌ {str(e)}")
                return
        else:
            now = datetime.now()
            start_date = now.replace(day=1, hour=0, minute=0, second=0)
            end_date = now.replace(day=calendar.monthrange(now.year, now.month)[1],
                                    hour=23, minute=59, second=59)

        # 获取数据（只查询一次）
        balances = session.query(Balance).filter_by(customer_name=customer).all()
        txs = session.query(Transaction).filter(
            Transaction.customer_name == customer,
            Transaction.timestamp.between(start_date, end_date),
            Transaction.transaction_type.in_(["buy", "sell"])
        ).all()
        adjs = session.query(Adjustment).filter(
            Adjustment.customer_name == customer,
            Adjustment.timestamp.between(start_date, end_date)
        ).all()

        # 查询所有交易记录（包括支付记录），并按时间排序
        all_records = session.query(Transaction).filter(
            Transaction.customer_name == customer,
            Transaction.timestamp.between(start_date, end_date),
            or_(
                Transaction.status != 'canceled',
                Transaction.status.is_(None)
            )
        ).all()
        sorted_records = sorted(all_records, key=lambda x: x.timestamp)

        # 获取客户当前所有余额中的货币
        balance_currencies = {b.currency.upper() for b in session.query(Balance).filter_by(customer_name=customer).all()}
        tx_currencies = set()
        for tx in txs:
            if tx.base_currency.upper() != '-':
                tx_currencies.add(tx.base_currency.upper())
            if tx.quote_currency.upper() != '-':
                tx_currencies.add(tx.quote_currency.upper())
        all_currencies = balance_currencies.union(tx_currencies)
        sorted_currencies = sorted(list(all_currencies))

        # ===== 步骤2：动态余额计算系统 =====
        currency_balances = defaultdict(float)
        processed_records = []

        # 查询开始日期前的交易和调整记录以计算期初余额
        initial_txs = session.query(Transaction).filter(
            Transaction.customer_name == customer,
            Transaction.timestamp < start_date
        ).all()

        initial_adjs = session.query(Adjustment).filter(
            Adjustment.customer_name == customer,
            Adjustment.timestamp < start_date
        ).all()

        # ===== 步骤1：查询当前余额（以数据库中 Balance 表为准，确保与 Telegram 响应一致） =====
        current_balances = {
            b.currency.upper(): round(b.amount, 2)
            for b in session.query(Balance).filter_by(customer_name=customer).all()
        }

        # ===== 步骤2：计算报告期间内的净变化 =====
        # 注意：这里所有金额都做了 round(..., 2) 处理，以保持与 update_balance() 一致
        # 查询报告期间内所有交易记录（包括买入、卖出和支付记录）
        period_txs = session.query(Transaction).filter(
            Transaction.customer_name == customer,
            Transaction.timestamp.between(start_date, end_date)
        ).all()

        net_changes = defaultdict(float)
        for tx in period_txs:
            if tx.transaction_type == 'buy':
                base = tx.base_currency.upper()
                quote = tx.quote_currency.upper()
                # 计算报价金额时同样用 round 保持一致性
                total_quote = round(tx.amount / tx.rate, 2) if tx.operator == '/' else round(tx.amount * tx.rate, 2)
                # 买入：客户获得基础币（增加），支付报价币（减少）
                net_changes[base] += round(tx.amount, 2)
                net_changes[quote] -= total_quote
            elif tx.transaction_type == 'sell':
                base = tx.base_currency.upper()
                quote = tx.quote_currency.upper()
                total_quote = round(tx.amount / tx.rate, 2) if tx.operator == '/' else round(tx.amount * tx.rate, 2)
                # 卖出：客户支付基础币（减少），获得报价币（增加）
                net_changes[base] -= round(tx.amount, 2)
                net_changes[quote] += total_quote
            elif tx.transaction_type == 'payment':
                # 对于支付记录，根据支付类型调整
                if tx.sub_type == '客户支付':
                    curr = tx.quote_currency.upper()
                    net_changes[curr] += round(tx.settled_in, 2)
                elif tx.sub_type == '公司支付':
                    curr = tx.base_currency.upper()
                    net_changes[curr] -= round(tx.settled_out, 2)

        # ===== 步骤3：推算期初余额 =====
        initial_balances = {}
        for curr, curr_balance in current_balances.items():
            # 期初余额 = 当前余额 - 期间净变化
            initial_balances[curr] = round(curr_balance - net_changes.get(curr, 0.0), 2)

        # 示例：构造一个"期初余额"记录行，用于Excel报表显示
        initial_record = {
            "日期": start_date.strftime('%Y-%m-%d'),
            "订单号": "期初余额",
            "类型": "期初",
            "交易对": "-",
            "数量": "-",
            "总额": "-",
            "汇率": "-",
            "进度": "-",
            "状态": "-"
        }
        for curr in sorted(initial_balances.keys()):
            value = initial_balances[curr]
            # 只显示不为 0 的币种（这里判断绝对值小于 0.01 视为 0）
            if abs(value) < 0.01:
                initial_record[f"{curr}余额"] = ""
            else:
                initial_record[f"{curr}余额"] = f"{value:+,.2f}"
        processed_records.append(initial_record)

        # 初始化currency_balances为期初余额
        currency_balances = defaultdict(float, initial_balances)

        # 然后处理期间内的交易记录
        for tx in sorted_records:
            # 如果是支付记录，单独处理
            if tx.transaction_type == 'payment':
                # 简单规则：客户支付记作加（+），公司支付记作减（-）
                record = {
                    "日期": tx.timestamp.strftime('%Y-%m-%d'),
                    "订单号": "",
                    "类型": "",
                    "交易对": "-",
                    "数量": "-",
                    "总额": "-",
                    "汇率": "-",
                    "进度": "-",
                    "状态": "-"
                }
                # 初始化所有币种余额列为空
                for curr in sorted_currencies:
                    record[f"{curr}余额"] = ""
                if tx.sub_type == '客户支付':
                    # 假设客户支付时，支付币种存于 quote_currency
                    curr = tx.quote_currency.upper()
                    record[f"{curr}余额"] = f"+{tx.amount:,.2f}"
                    record["订单号"] = f"客户支付({tx.amount:,.2f} {curr})"
                    currency_balances[curr] += tx.amount
                elif tx.sub_type == '公司支付':
                    # 公司支付时，支付币种存于 base_currency
                    curr = tx.base_currency.upper()
                    record[f"{curr}余额"] = f"-{tx.amount:,.2f}"
                    record["订单号"] = f"公司支付({tx.amount:,.2f} {curr})"
                    currency_balances[curr] -= tx.amount
                else:
                    # 其它支付记录，留空或按需要处理
                    pass
                processed_records.append(record)
                continue  # 跳过后续处理

            # 对于普通交易记录（买入或卖出）保持原逻辑
            base_curr = tx.base_currency.upper()
            quote_curr = tx.quote_currency.upper()
            if tx.rate == 0:
                total_quote = 0
            else:
                total_quote = tx.amount / tx.rate if tx.operator == '/' else tx.amount * tx.rate

            exchange_rate = round(tx.amount / total_quote, 6) if total_quote else 0
            progress = "0.0%"
            if tx.amount != 0 and total_quote != 0:
                if tx.transaction_type == 'buy':
                    progress_value = min(tx.settled_in / total_quote, tx.settled_out / tx.amount)
                else:
                    progress_value = min(tx.settled_in / tx.amount, tx.settled_out / total_quote)
                progress = f"{progress_value * 100:.1f}%"

            record = {
                "日期": tx.timestamp.strftime('%Y-%m-%d'),
                "订单号": tx.order_id,
                "类型": "买入" if tx.transaction_type == 'buy' else "卖出",
                "交易对": f"{base_curr}/{quote_curr}",
                "数量": f"{tx.amount:,.2f} {base_curr}",
                "总额": f"{total_quote:,.2f} {quote_curr}",
                "汇率": f"1 {quote_curr} = {exchange_rate:.6f} {base_curr}",
                "进度": progress,
            }
            for curr in sorted_currencies:
                value = currency_balances.get(curr, 0)
                if abs(value) < 0.01:
                    record[f"{curr}余额"] = ""
                else:
                    record[f"{curr}余额"] = f"{value:+,.2f}"

            # 根据交易类型更新余额：
            # 买入：客户支付（报价币）减少，获得基础币增加
            # 卖出：客户支付（基础币）减少，获得报价币增加
            if tx.transaction_type == 'buy':
                currency_balances[quote_curr] -= total_quote
                currency_balances[base_curr] += tx.amount
            else:
                currency_balances[base_curr] -= tx.amount
                currency_balances[quote_curr] += total_quote

            # 将当前累计余额写入记录（简单显示累计值即可）
            for curr in sorted_currencies:
                value = currency_balances.get(curr, 0)
                if abs(value) < 0.01:
                    record[f"{curr}余额"] = ""
                else:
                    record[f"{curr}余额"] = f"{value:+,.2f}"
            processed_records.append(record)

        # ===== 步骤3：生成最终余额行 =====
        final_balance = {
            "日期": "当前余额",
            "订单号": "当前余额",
            "类型": "余额汇总",
            "交易对": "",
            "数量": "",
            "总额": "",
            "汇率": "",
            "进度": "",
        }
        for curr in sorted_currencies:
            final_balance[f"{curr}余额"] = f"{currency_balances.get(curr, 0):+,.2f}"
        processed_records.append(final_balance)

        # ===== 步骤4：输出报表 =====
        if excel_mode:
            base_columns = ['日期', '订单号', '类型', '交易对', '数量', '总额', '汇率', '进度']
            dynamic_columns = [f"{c}余额" for c in sorted_currencies]
            df = pd.DataFrame(processed_records)[base_columns + dynamic_columns]
            excel_buffer = generate_excel_buffer({'交易明细与余额': df}, ["交易明细与余额"])
            await update.message.reply_document(
                document=excel_buffer,
                filename=f"客户对账单_{customer}_动态货币版.xlsx",
                caption=f"📊 {customer} 对账单（支持{len(sorted_currencies)}种货币）"
            )
            return
        elif image_mode:
            # Generate and send image
            img_buffer = await generate_statement_image(processed_records, customer, start_date, end_date)
            await update.message.reply_photo(
                photo=img_buffer,
                caption=f"📊 {customer} 对账单\n{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
            )
            return

        # 生成文本报告
        report = [
            f"📑 客户对账单 - {customer}",
            f"日期范围: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
            f"生成时间: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            "━━━━━━━━━━━━━━━━━━"
        ]

        balance_section = ["📊 当前余额:"]
        if balances:
            balance_section += [f"• {b.currency}: {b.amount:+,.2f}" for b in balances]
        report.extend(balance_section)

        tx_section = ["\n💵 交易记录:"]
        if txs:
            for tx in txs:
                if tx.operator == '/':
                    total_quote = tx.amount / tx.rate
                else:
                    total_quote = tx.amount * tx.rate

                if tx.transaction_type == 'buy':
                    settled_base = tx.settled_out
                    settled_quote = tx.settled_in
                else:
                    settled_base = tx.settled_in
                    settled_quote = tx.settled_out

                base_progress = settled_base / tx.amount if tx.amount != 0 else 0
                quote_progress = settled_quote / total_quote if total_quote != 0 else 0
                base_done = int(settled_base) >= int(tx.amount)
                quote_done = int(settled_quote) >= int(total_quote)
                status = "已完成" if base_done and quote_done else "进行中"

                tx_section.append(
                    f"▫️ {tx.timestamp.strftime('%d/%m %H:%M')} {tx.order_id}\n"
                    f"{'买入' if tx.transaction_type == 'buy' else '卖出'} {tx.amount:,.2f} {tx.base_currency} @ {tx.rate:.4f}\n"
                    f"├─ 已结基础货币: {settled_base:,.2f}/{tx.amount:,.2f} {tx.base_currency} ({base_progress:.1%})\n"
                    f"├─ 已结报价货币: {settled_quote:,.2f}/{total_quote:,.2f} {tx.quote_currency} ({quote_progress:.1%})\n"
                    f"└─ 状态: {status}"
                )
        else:
            tx_section.append("无交易记录")
        report.extend(tx_section)

        adj_section = ["\n📝 调整记录:"]
        if adjs:
            for adj in adjs:
                adj_section.append(
                    f"{adj.timestamp.strftime('%d/%m %H:%M')}\n"
                    f"{adj.currency}: {adj.amount:+,.2f} - {adj.note}"
                )
        else:
            adj_section.append("无调整记录")
        report.extend(adj_section)

        full_report = "\n".join(report)
        for i in range(0, len(full_report), 4000):
            await update.message.reply_text(full_report[i:i+4000])
    except Exception as e:
        logger.error(f"对账单生成失败: {str(e)}")
        await update.message.reply_text("❌ 生成失败")
    finally:
        Session.remove()
        
async def average_cost(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """查询公司当前持有的USDT和MYR平均成本"""
    session = Session()
    try:
        usdt = session.query(USDTAverageCost).first()
        myr = session.query(MYRAverageCost).first()
        
        usdt_avg = usdt.average_cost if usdt else 0.0
        myr_avg = myr.average_cost if myr else 0.0
        
        response = (
            "📊 *公司持仓均价报告*\n"
            "━━━━━━━━━━━━━━━━━━\n"
            f"• USDT 平均成本价: {usdt_avg:.4f} MYR/USDT\n"
            f"   ▸ 公司持有 {usdt.total_usdt if usdt else 0:,.2f} USDT\n"
            f"   ▸ 累计消耗 {usdt.total_myr_spent if usdt else 0:,.2f} MYR\n\n"
            f"• MYR 平均成本价: {myr_avg:.4f} USDT/MYR\n"
            f"   ▸ 公司持有 {myr.total_myr if myr else 0:,.2f} MYR\n"
            f"   ▸ 累计消耗 {myr.total_usdt_spent if myr else 0:,.2f} USDT\n"
            "━━━━━━━━━━━━━━━━━━\n"
            "注：仅统计涉及MYR/USDT货币对的交易"
        )
        await update.message.reply_text(response, parse_mode="Markdown")
    except Exception as e:
        logger.error(f"均价查询失败: {str(e)}")
        await update.message.reply_text("❌ 查询失败，请检查日志")
    finally:
        Session.remove()

async def generate_statement_image(records, customer, start_date, end_date):
    """Generate an image of the customer statement with Excel-like styling"""

    # ===== 1. 先定义基础参数 =====
    padding = 30
    row_height = 55
    header_height = 140
    
    # 颜色定义
    header_color = '#D3D3D3'
    border_color = '#000000'
    alt_row_color = '#FFFFFF'
    initial_balance_color = '#B4C6E7'
    final_balance_color = '#FFFF00'
    customer_payment_color = '#C6EFCE'
    company_payment_color = '#FFC7CE'
    positive_color = '#008000'
    negative_color = '#FF0000'
    
    # 货币列（如果 records 为空，下面这行需要额外判断）
    currency_cols = [col for col in records[0].keys() if col.endswith('余额')]
    
    # 初步给个“默认列宽”或者“最小列宽”
    col_widths = {
        "日期": 100,
        "订单号": 180,
        "类型": 80,
        "交易对": 120,
        "数量": 150,
        "总额": 150,
        "汇率": 260,
        "进度": 100,
    }
    for curr_col in currency_cols:
        col_widths[curr_col] = 120
    
    # ===== 2. 定义字体（加大字号） =====
    try:
        title_font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 28)
        header_font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 20)
        font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 18)
        bold_font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 20, index=1)
    except:
        # Fallback
        title_font = ImageFont.load_default()
        header_font = title_font
        font = title_font
        bold_font = title_font
    
    # ===== 3. 创建一个“虚拟画布”用来测量文本宽度 =====
    dummy_img = Image.new('RGB', (1, 1), 'white')
    dummy_draw = ImageDraw.Draw(dummy_img)
    
    # ===== 4. 动态计算列宽 =====
    for col in col_widths.keys():
        # 先测量表头
        header_bbox = dummy_draw.textbbox((0, 0), col, font=header_font)
        needed_header_width = header_bbox[2] - header_bbox[0]
        
        # 再测量该列数据中最宽的那一个
        max_data_width = 0
        for record in records:
            order_id = str(record.get('订单号', '')).strip()
            is_final_row = (order_id == '当前余额')
            # 根据你的逻辑决定用哪个字体
            if order_id == '期初余额':
                used_font = font
            elif is_final_row:
                used_font = bold_font
            elif '客户支付' in order_id or '公司支付' in order_id:
                used_font = font
            else:
                used_font = font
            
            text = str(record.get(col, ''))
            data_bbox = dummy_draw.textbbox((0, 0), text, font=used_font)
            data_width = data_bbox[2] - data_bbox[0]
            if data_width > max_data_width:
                max_data_width = data_width
        
        # 最终列宽 = max(表头宽, 最大数据宽) + 适当的安全边距
        final_col_width = max(needed_header_width, max_data_width) + 20
        col_widths[col] = int(final_col_width)
    
    # ===== 5. 计算最终图像大小，再创建真正的图像 =====
    width = sum(col_widths.values()) + padding * 2
    height = header_height + (len(records) + 1) * row_height + padding * 2
    
    img = Image.new('RGB', (width, height), 'white')
    draw = ImageDraw.Draw(img)
    
    # ===== 6. 绘制标题、表头等 =====
    title = f"客户对账单 - {customer}"
    date_range = f"日期范围: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
    draw.text((padding, padding), title, font=title_font, fill='black')
    draw.text((padding, padding + 40), date_range, font=header_font, fill='black')
    
    # 表头起始位置
    y = header_height
    x = padding
    
    def draw_cell(x, y, w, text, font, fill_color=None, text_color='black', align='left', is_final=False):
        # 画背景
        if fill_color:
            draw.rectangle([x, y, x + w, y + row_height], fill=fill_color)
        
        # 边框
        border_width = 2 if is_final else 1
        draw.rectangle([x, y, x + w, y + row_height], outline=border_color, width=border_width)
        
        # 计算文本位置
        text_bbox = draw.textbbox((0, 0), text, font=font)
        text_width = text_bbox[2] - text_bbox[0]
        text_height = text_bbox[3] - text_bbox[1]
        
        if align == 'center':
            text_x = x + (w - text_width) // 2
        elif align == 'right':
            text_x = x + w - text_width - 5
        else:
            text_x = x + 5
        
        draw.text((text_x, y + (row_height - text_height) // 2),
                  text, font=font, fill=text_color)
    
    # 绘制表头
    for col, w in col_widths.items():
        draw_cell(x, y, w, col, header_font, fill_color=header_color, align='center')
        x += w
    
    # ===== 7. 绘制数据行 =====
    for idx, record in enumerate(records):
        y += row_height
        x = padding
        
        order_id = str(record.get('订单号', '')).strip()
        is_final_row = (order_id == '当前余额')
        
        if order_id == '期初余额':
            row_color = initial_balance_color
            font_to_use = font
        elif is_final_row:
            row_color = final_balance_color
            font_to_use = bold_font
        elif '客户支付' in order_id:
            row_color = customer_payment_color
            font_to_use = font
        elif '公司支付' in order_id:
            row_color = company_payment_color
            font_to_use = font
        else:
            row_color = alt_row_color if idx % 2 == 0 else 'white'
            font_to_use = font
        
        for col, w in col_widths.items():
            value = str(record.get(col, ''))
            
            text_color = 'black'
            # 如果是余额列，则根据正负决定颜色
            if col.endswith('余额') and value:
                try:
                    num_value = float(value.replace(',', '').replace('+', ''))
                    text_color = positive_color if num_value > 0 else negative_color
                except ValueError:
                    pass
            
            align = 'right' if col in ['数量', '总额', '汇率'] or col.endswith('余额') else 'left'
            
            draw_cell(x, y, w, value, font_to_use, row_color, text_color, align, is_final_row)
            x += w
    
    # ===== 8. 转为字节流输出 =====
    img_byte_array = io.BytesIO()
    img.save(img_byte_array, format='PNG', quality=95)
    img_byte_array.seek(0)
    
    return img_byte_array

def generate_cashflow_image_side_summary(
    payments_data,
    start_date,
    end_date,
    totals_customer,
    totals_company
):
    """
    生成支付流水图片，将"汇总"信息放在右侧的独立面板。
    - 左侧：表头 + 数据行
    - 右侧：汇总（客户支付、公司支付、净流量）
    - 字体加大，布局更宽松、更易读
    """
    import io
    from PIL import Image, ImageDraw, ImageFont

    # ========== 1) 样式与字体 ==========
    # 整体留白
    padding = 40
    # 表格行高
    row_height = 65
    # 标题区高度
    header_height = 140
    # 表格线条
    border_color = '#000000'
    line_width = 2

    # 颜色
    header_fill = '#E8E8E8'         # 表头背景：浅灰
    customer_fill = '#D4F4DD'       # 客户支付行：浅绿
    company_fill = '#FDDADA'        # 公司支付行：浅红
    white_fill = '#FFFFFF'
    summary_fill = '#FFFF00'        # 汇总区标题行：淡黄

    # 尝试加载更大的字体
    try:
        title_font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 40)
        header_font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 28)
        font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 24)
        bold_font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 24, index=1)
    except:
        title_font = ImageFont.load_default()
        header_font = title_font
        font = title_font
        bold_font = title_font

    # ========== 2) 定义表格列、初始列宽、对齐方式 ==========
    columns = [
        {"key": "时间",  "title": "时间",   "width": 100, "align": "center"},
        {"key": "订单号","title": "订单号", "width": 240, "align": "left"},
        {"key": "客户",  "title": "客户",   "width": 120, "align": "center"},
        {"key": "类型",  "title": "类型",   "width": 120, "align": "center"},
        {"key": "金额",  "title": "金额",   "width": 140, "align": "right"},
        {"key": "币种",  "title": "币种",   "width": 90,  "align": "right"},
    ]

    dummy_img = Image.new('RGB', (10, 10), 'white')
    dummy_draw = ImageDraw.Draw(dummy_img)

    def measure_text(txt, fnt):
        bbox = dummy_draw.textbbox((0, 0), txt, font=fnt)
        return (bbox[2] - bbox[0], bbox[3] - bbox[1])

    # ========== 3) 动态调整列宽 ==========
    # 先表头
    for col in columns:
        w, h = measure_text(col["title"], header_font)
        col["width"] = max(col["width"], w + 30)

    # 再数据
    for row in payments_data:
        for col in columns:
            cell_text = row[col["key"]]
            w, h = measure_text(cell_text, font)
            if w + 30 > col["width"]:
                col["width"] = w + 30

    table_width = sum(col["width"] for col in columns)
    data_rows = len(payments_data)

    # ========== 4) 右侧汇总面板计算 ==========
    # 先把客户支付、公司支付、净流量的行数算一下
    ccy_customer = set(totals_customer.keys())
    ccy_company = set(totals_company.keys())
    ccy_all = ccy_customer.union(ccy_company)

    # 汇总行包含：
    # - 标题 "Summary"
    # - "客户支付总额" + N 行
    # - "公司支付总额" + M 行
    # - "净流量(Net Flow)" + K 行
    #   (K = 并集币种数量)
    summary_lines = 1 + (1 + len(ccy_customer)) + (1 + len(ccy_company)) + (1 + len(ccy_all))
    summary_width = 360  # 固定右侧面板宽度，可根据需要调整
    summary_height = summary_lines * row_height

    # ========== 5) 计算画布整体宽高 ==========
    # 左表宽 + 右面板宽 + 3*padding (左、中、右)
    img_width = table_width + summary_width + padding * 3
    # 表格区高度：标题区 + (表头+数据行)*row_height
    table_height = header_height + (1 + data_rows)*row_height
    # 总体高度 = max(表格区高度, 标题区 + summary_height) + padding
    content_height = max(table_height, header_height + summary_height)
    img_height = content_height + padding

    # ========== 6) 创建画布、绘制标题 ==========
    img = Image.new('RGB', (img_width, img_height), color='white')
    draw = ImageDraw.Draw(img)

    title_text = "支付流水报告"
    date_range_text = f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"

    # 标题
    draw.text((padding, padding), title_text, font=title_font, fill='black')
    draw.text((padding, padding + 60), f"日期范围: {date_range_text}", font=header_font, fill='black')

    # 表格起始坐标
    table_x = padding
    table_y = padding + header_height

    # ========== 7) 绘制表头 ==========
    x = table_x
    y = table_y
    for col in columns:
        w = col["width"]
        draw.rectangle([x, y, x + w, y + row_height], fill=header_fill, outline=border_color, width=line_width)
        col_title = col["title"]
        txt_w, txt_h = measure_text(col_title, header_font)
        if col["align"] == "center":
            text_x = x + (w - txt_w)//2
        elif col["align"] == "right":
            text_x = x + w - txt_w - 10
        else:
            text_x = x + 10
        text_y = y + (row_height - txt_h)//2
        draw.text((text_x, text_y), col_title, font=header_font, fill='black')
        x += w
    y += row_height

    # ========== 8) 绘制数据行 ==========
    for row in payments_data:
        row_type = row["类型"]
        if row_type == "客户支付":
            bg_color = customer_fill
        elif row_type == "公司支付":
            bg_color = company_fill
        else:
            bg_color = white_fill

        x = table_x
        for col in columns:
            w = col["width"]
            cell_text = row[col["key"]]
            draw.rectangle([x, y, x + w, y + row_height], fill=bg_color, outline=border_color, width=line_width)
            txt_w, txt_h = measure_text(cell_text, font)
            if col["align"] == "center":
                text_x = x + (w - txt_w)//2
            elif col["align"] == "right":
                text_x = x + w - txt_w - 10
            else:
                text_x = x + 10
            text_y = y + (row_height - txt_h)//2
            draw.text((text_x, text_y), cell_text, font=font, fill='black')
            x += w
        y += row_height

    # ========== 9) 绘制右侧汇总面板 ==========
    # 面板左上角
    summary_x = table_x + table_width + padding
    summary_y = padding + header_height

    # 定义一个函数，用来画一行
    def draw_summary_line(draw, text, x, y, width, fill_color=white_fill, fnt=font, bold=False):
        draw.rectangle([x, y, x + width, y + row_height], fill=fill_color, outline=border_color, width=line_width)
        txt_w, txt_h = measure_text(text, fnt)
        # 左对齐
        draw.text((x + 10, y + (row_height - txt_h)//2), text, font=fnt, fill='black')

    current_y = summary_y

    # (1) 标题 "Summary"
    draw_summary_line(draw, "Summary", summary_x, current_y, summary_width, fill_color=summary_fill, fnt=bold_font, bold=True)
    current_y += row_height

    # (2) 客户支付总额
    draw_summary_line(draw, "客户支付总额：", summary_x, current_y, summary_width, fill_color=summary_fill, fnt=bold_font)
    current_y += row_height
    for ccy in sorted(totals_customer.keys()):
        amt = totals_customer[ccy]
        line_str = f"{ccy}: {amt:,.2f}"
        draw_summary_line(draw, line_str, summary_x, current_y, summary_width, fill_color=white_fill, fnt=font)
        current_y += row_height

    # (3) 公司支付总额
    draw_summary_line(draw, "公司支付总额：", summary_x, current_y, summary_width, fill_color=summary_fill, fnt=bold_font)
    current_y += row_height
    for ccy in sorted(totals_company.keys()):
        amt = totals_company[ccy]
        line_str = f"{ccy}: {amt:,.2f}"
        draw_summary_line(draw, line_str, summary_x, current_y, summary_width, fill_color=white_fill, fnt=font)
        current_y += row_height

    # (4) 净流量 (Net Flow)
    draw_summary_line(draw, "净流量 (Net Flow)：", summary_x, current_y, summary_width, fill_color=summary_fill, fnt=bold_font)
    current_y += row_height
    all_ccy = set(totals_customer.keys()) | set(totals_company.keys())
    for ccy in sorted(all_ccy):
        in_amt = totals_customer.get(ccy, 0.0)
        out_amt = totals_company.get(ccy, 0.0)
        net = in_amt - out_amt
        sign_str = "+" if net > 0 else ""
        line_str = f"{ccy}: {sign_str}{net:,.2f}"
        draw_summary_line(draw, line_str, summary_x, current_y, summary_width, fill_color=white_fill, fnt=font)
        current_y += row_height

    # ========== 10) 导出到 BytesIO ==========
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    return buffer


async def cash_flow_report_side_summary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /cashflow 命令的实现，将汇总面板放到右侧。
    用法：
      /cashflow
      /cashflow 02/03/2025
      /cashflow 01/03/2025-05/03/2025
    """
    session = Session()
    try:
        args = context.args
        if not args:
            start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date + timedelta(days=1) - timedelta(microseconds=1)
        else:
            date_str = " ".join(args)
            if '-' in date_str:
                # 视为日期范围
                try:
                    sd, ed = parse_date_range(date_str)
                    start_date = sd
                    end_date = ed
                except ValueError as e:
                    await update.message.reply_text(f"❌ {str(e)}")
                    return
            else:
                # 单个日期
                try:
                    single_date = datetime.strptime(date_str.strip(), '%d/%m/%Y')
                    start_date = single_date.replace(hour=0, minute=0, second=0, microsecond=0)
                    end_date = single_date.replace(hour=23, minute=59, second=59, microsecond=999999)
                except:
                    await update.message.reply_text("❌ 日期格式错误，请使用 DD/MM/YYYY 或 DD/MM/YYYY-DD/MM/YYYY")
                    return

        payments = session.query(Transaction).filter(
            Transaction.transaction_type == 'payment',
            Transaction.timestamp >= start_date,
            Transaction.timestamp <= end_date,
            or_(Transaction.status != 'canceled', Transaction.status.is_(None))
        ).order_by(Transaction.timestamp.asc()).all()

        if not payments:
            await update.message.reply_text("指定日期内无支付记录")
            return

        # 整理数据
        payments_data = []
        totals_customer = defaultdict(float)
        totals_company = defaultdict(float)

        for p in payments:
            time_str = p.timestamp.strftime('%H:%M')
            sub_type = p.sub_type or "-"
            if sub_type == "客户支付":
                amount = p.settled_in
                currency = p.quote_currency
            elif sub_type == "公司支付":
                amount = p.settled_out
                currency = p.base_currency
            else:
                amount = p.amount
                currency = p.base_currency

            row = {
                "时间": time_str,
                "订单号": p.order_id,
                "客户": p.customer_name or "",
                "类型": sub_type,
                "金额": f"{amount:,.2f}",
                "币种": currency or ""
            }
            payments_data.append(row)

            if sub_type == "客户支付":
                totals_customer[currency] += amount
            elif sub_type == "公司支付":
                totals_company[currency] += amount

        # 调用上面定义的函数，生成右侧汇总面板版图片
        img_buffer = generate_cashflow_image_side_summary(
            payments_data,
            start_date,
            end_date,
            totals_customer,
            totals_company
        )
        await update.message.reply_photo(photo=img_buffer, caption="📊 支付流水报告（右侧汇总）")

    except Exception as e:
        logger.error(f"生成支付流水报告失败: {str(e)}", exc_info=True)
        await update.message.reply_text("❌ 生成支付流水报告失败，请查看日志")
    finally:
        Session.remove()


# ================== 机器人命令注册 ==================
def main():
    run_migrations()  # 新增此行
    setup_logging()
    application = ApplicationBuilder().token("YOUR_BOT_TOKEN").build()
    
    handlers = [
        CommandHandler('start', lambda u, c: u.message.reply_text(
            "🤖 *会计机器人*\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "📚 可用命令：\n\n"
            "💼 *账户管理*\n"
            "▫️ `/balance [客户]` 查询余额 📊\n"
            "▫️ `/debts [客户]` 查看欠款明细 🧾\n"
            "▫️ `/adjust [客户] [货币] [±金额] [备注]` 调整余额 ⚖️\n\n"
            "▫️ `/delete_customer [客户名]` 删除客户及其所有数据 ⚠️\n\n"
            "💸 *交易操作*\n"
            "▫️ `客户A 买 10000USD /4.42 MYR` 创建交易\n"
            "▫️ `/received [客户] [金额+币种]` 登记客户付款\n"
            "▫️ `/paid [客户] [金额+币种]` 登记向客户付款\n"
            "▫️ `/cancel [订单号]` 撤销未结算交易\n\n"
            "📈 *财务报告*\n"
            "▫️ `/pnl [日期范围] [excel]` 盈亏报告 📉\n"
            "▫️ `/report [日期范围] [excel]` 交易明细 📋\n"
            "▫️ `/creport [客户] [日期范围] [excel/image]` 客户对账单 📑\n"
            "▫️ `/expense [金额+币种] [用途]` 记录支出 💸\n"
            "▫️ `/expenses` 支出记录 🧮\n\n"
            "▫️ `/average` 计算公司货币的持仓均价 📈\n"
            "▫️ `/cashflow` 生成今日支付流水报告（图片）\n\n"
            "💡 *使用提示*\n"
            "🔸 日期格式：`DD/MM/YYYY-DD/MM/YYYY`\n"
            "🔸 添加 `excel` 参数获取表格文件 📤\n"
            "🔸 示例：`/pnl 01/01/2025-31/03/2025 excel`"
        )),
        CommandHandler('balance', balance),
        CommandHandler('debts', list_debts),
        CommandHandler('adjust', adjust_balance),
        CommandHandler('received', handle_received),
        CommandHandler('paid', handle_paid),
        CommandHandler('cancel', cancel_order),
        CommandHandler('cancel_payment', cancel_payment),
        CommandHandler('pnl', pnl_report),
        CommandHandler('detailed_pnl', detailed_pnl_report_cmd),
        CommandHandler('average', average_cost),
        CommandHandler('expense', add_expense),
        CommandHandler('expenses', list_expenses),
        CommandHandler('expensesimg', list_expenses_image),
        CommandHandler('creport', customer_statement),
        CommandHandler('report', lambda u, c: generate_detailed_report(u, c, 'daily')),
        CommandHandler('delete_customer', delete_customer),
        CommandHandler('cashflow', cash_flow_report_side_summary),
        MessageHandler(filters.TEXT & ~filters.COMMAND, handle_transaction)
    ]
    
    application.add_handlers(handlers)
    logger.info("机器人启动成功")
    application.run_polling()

if __name__ == '__main__':
    main()


